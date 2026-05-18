from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as  G

wb = Workbook()

# ── helpers ─────────────────────────────────────────────────────────────────
def tb():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def H(ws, r, c, v, bg="1F4E79", fg="FFFFFF", sz=10, bold=True, wrap=True, align="center"):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    x.border = tb()
    return x

def C(ws, r, c, v, bold=False, color="111111", sz=9, wrap=True, bg=None, align="left", italic=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name="Arial", bold=bold, color=color, size=sz, italic=italic)
    x.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg: x.fill = PatternFill("solid", fgColor=bg)
    x.border = tb()
    return x

def T(ws, text, bg, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    x = ws.cell(row=row, column=1, value=text)
    x.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32

def W(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[G(i)].width = w

def BD(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = tb()

ALT = ["D6EAF8", "EAFAF1", "FEF9E7", "FDEDEC", "F4ECF7", "E8F8F5", "FFF3CD", "EBF5FB", "D5F5E3", "FDE8D8"]

# ── shared text blocks ───────────────────────────────────────────────────────
SIG = """Sincerely,
Abd El Madjid Kahoul
MSc Software Engineering & Intelligent Systems — University of Constantine 2, Algeria
abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17
LinkedIn: kahoul-abd-el-madjid
Scopus Q2 Paper: DOI 10.18080/jtde.v13n4.1363
Scopus Paper 2: DOI 10.18080/jtde.v14n1.142"""

SENTRY_BRIEF = """My current PhD research proposal, SENTRY, addresses a critical unsolved problem: multi-agent LLM systems deployed in production fail silently — reasoning loops, tool-call errors, and inter-agent delegation drift produce no observable error signal. SENTRY is a non-invasive runtime monitoring framework that hooks into agent event buses (LangGraph/AutoGen) without modifying agent code, constructs dynamic inter-agent interaction graphs, and applies a Graph Attention Network (GAT) to detect structural behavioural anomalies. A graduated intervention engine (Monitor → Pause → Rollback → Handoff) responds proportionally, and SHAP/GNNExplainer-based explainability produces operator-readable justifications. This bridges my published work on anomaly detection and XAI in operational systems with the emerging field of trustworthy multi-agent AI."""

PAPERS = """My two Scopus publications provide the methodological foundation:
(1) First-author: "Hybrid Deep Learning Ensemble with Dynamic Fusion for Reliable Anomaly Detection in Operational LTE RANs" — JTDE Scopus Q2, Vol.13 No.4, 2025. DOI: 10.18080/jtde.v13n4.1363 — hybrid DL (autoencoders + LSTMs + transformers + ensemble fusion) for anomaly detection in real production telecom data from Ooredoo Algeria.
(2) Co-author: "RAVA: A Human-in-the-Loop Hybrid Platform for Explainable Anomaly Detection and Diagnosis in LTE RANs" — JTDE, Vol.14 No.1, 2026. DOI: 10.18080/jtde.v14n1.142 — SHAP-based XAI in an interactive operator-facing diagnosis workflow."""

COND = """Regarding English proficiency: I am currently preparing for IELTS. I am available for a research interview at any time to demonstrate my English proficiency and technical depth, and I would warmly welcome a conditional acceptance arrangement — providing the certificate before enrollment."""

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER ACTION PLAN
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "🗺️ MASTER PLAN"
ws1.sheet_view.showGridLines = False
T(ws1, "MASTER ACTION PLAN — Step-by-Step, Week-by-Week | What to Send · To Whom · When · How", "1A5276", 5)
for i, h in enumerate(["WHEN", "EXACT ACTION (what to do)", "WHO TO CONTACT (name + email)", "WHAT TO SEND (exact documents)", "STATUS"], 1):
    H(ws1, 2, i, h, bg="2E86C1", sz=9)
ws1.row_dimensions[2].height = 28

plan = [
    ("WEEK 1\nMay 12–16, 2026\n\n🔴 URGENT:\nITALY PNRR opens\nJAPAN MEXT opens\nEUREXESS has\nopen positions NOW",
     """STEP 1 — Email Italian professors (PNRR calls opening NOW):
→ Prof. Chiasserini (PoliTo) — explicitly named in your SENTRY proposal
→ Prof. Montali (Univ. Bolzano) — process mining / runtime monitoring
→ Prof. Ardagna (PoliMi) — distributed AI systems

STEP 2 — Contact Japanese Embassy Algiers for MEXT:
→ Address: 25 Chemin Gadouche, Hydra, Algiers
→ Phone: +213 21 54 04 08
→ Ask: "I want to apply for the 2026 MEXT research scholarship"
→ Deadline for embassy route: May–June 2026

STEP 3 — Set up EURAXESS alert:
→ Go to euraxess.ec.europa.eu/jobs/search
→ Search: "trustworthy AI", "multi-agent systems", "graph neural networks"
→ Click "Create Alert" — new positions emailed to you automatically

STEP 4 — Prepare your document package:
→ Check: English translation of transcripts ✅ (you confirmed you have these)
→ Finalize 2-page CV with SENTRY proposal bullet points added
→ Prepare 3-page research proposal summary (extract from SENTRY PDF)""",
     "Prof. Chiasserini: chiasserini@polito.it\nProf. Montali: marco.montali@unibz.it\nProf. Ardagna: danilo.ardagna@polimi.it\n\nJapanese Embassy Algiers:\n+213 21 54 04 08\n\nEUREXESS: euraxess.ec.europa.eu",
     "• Email bodies: see Sheet ITALY for exact text\n• Attach: CV (2 pages)\n• Attach: SENTRY proposal PDF\n• No IELTS needed yet\n• No motivation letter at this stage",
     "⬜ TODO"),
    ("WEEK 2\nMay 19–23, 2026",
     """STEP 5 — Email Belgium/Netherlands/Austria professors (TOP PRIORITY):
→ Prof. De Turck (Ghent IDLab) — EXPLICITLY NAMED in your SENTRY proposal
→ Prof. Dustdar (TU Wien DSG) — EXPLICITLY NAMED in your SENTRY proposal
→ Prof. Dastani (Utrecht) — Europe's top multi-agent systems researcher
→ Prof. Fisher (Manchester) — agent verification expert

STEP 6 — Email France professors:
→ Prof. François (Inria RESIST) — in FRENCH (you have TCF B2)
→ He works on network anomaly detection monitoring (direct bridge from your papers to SENTRY)

STEP 7 — Check Italy PhD portals:
→ dottorato.polito.it — check if PNRR call is posted
→ phd.polimi.it — check if open
→ Apply formally if call is posted (see Sheet ITALY for exact steps)

STEP 8 — Prepare Malaysia MIS application:
→ Go to scholarship.moe.gov.my RIGHT NOW
→ Check if 2026 MIS cycle is open (typically March–May)
→ If open: email Malaysian professors immediately (see Sheet MALAYSIA)""",
     "Prof. De Turck: Filip.DeTurck@ugent.be\nProf. Dustdar: dustdar@dsg.tuwien.ac.at\nProf. Dastani: m.m.dastani@uu.nl\nProf. Fisher: Michael.Fisher@manchester.ac.uk\nProf. François: jerome.francois@inria.fr\n\nMalaysia MIS: scholarship.moe.gov.my",
     "• Emails: see Sheet EUROPE\n• Malaysia: see Sheet MALAYSIA\n• All emails: attach CV + SENTRY proposal\n• France email: write in French!",
     "⬜ TODO"),
    ("WEEK 3\nMay 26–30, 2026",
     """STEP 9 — Email Germany professors (DAAD target):
→ Prof. Müller (TU Berlin) — SHAP pioneer, directly cited in SENTRY
→ Prof. Thamsen (TU Berlin) — distributed systems anomaly detection
→ Ask both for DAAD supervisor letter (deadline November 2026)

STEP 10 — Email Japan professors:
→ Prof. Matsuo (UTokyo) — LLM research, agent systems
→ Prof. Sugiyama (RIKEN AIP) — anomaly detection + ML theory
→ Prof. Nakamura (NAIST) — multi-agent AI systems

STEP 11 — Follow up on Italy emails (if sent week 1):
→ Check replies from Chiasserini, Montali, Ardagna
→ If any reply: respond within 24 hours, schedule video call

STEP 12 — Begin MEXT application:
→ Download MEXT research plan form from Japanese embassy
→ Fill Section 1: personal background
→ Cite your 2 papers + SENTRY proposal in research plan""",
     "Prof. Müller: klaus-robert.mueller@tu-berlin.de\nProf. Thamsen: lauritz.thamsen@tu-berlin.de\nProf. Matsuo: matsuo@weblab.t.u-tokyo.ac.jp\nProf. Sugiyama: sugi@k.u-tokyo.ac.jp\nProf. Nakamura: s-nakamura@is.naist.jp",
     "• Germany emails: see Sheet EUROPE\n• Japan emails: see Sheet JAPAN\n• MEXT form: mext.go.jp/en\n• Research plan: 2 pages English",
     "⬜ TODO"),
    ("WEEK 4\nJun 2–6, 2026",
     """STEP 13 — Email Spain and Nordic professors:
→ Prof. Banchs (IMDEA) — ML for network systems (connects papers to SENTRY)
→ Prof. Latva-aho (Oulu, Finland) — 6G AI, trustworthy networks
→ Prof. Stiller (Zurich) — network management, AI audit

STEP 14 — Email Malaysian professors:
→ Prof. Mohd Sharifuddin (UTM) — multi-agent systems
→ Prof. Shahrul (UKM) — intelligent systems
→ If MIS is open: apply simultaneously at scholarship.moe.gov.my

STEP 15 — Interview preparation:
→ Prepare 20-minute research presentation covering:
   (a) Your 2 papers (5 min: what you built, real data, results)
   (b) SENTRY proposal (10 min: problem, architecture, experiments)
   (c) How they connect (5 min: XAI + anomaly detection expertise)
→ Practice in English: record yourself, review

STEP 16 — Book IELTS Academic exam:
→ Register at britishcouncil.dz or ielts.org
→ Target: June or July 2026 exam date
→ Target score: 6.5 (EU), 7.0 (Waterloo Canada)""",
     "Prof. Banchs: abanchs@it.uc3m.es\nProf. Latva-aho: matti.latva-aho@oulu.fi\nProf. Stiller: stiller@ifi.uzh.ch\nProf. Sharifuddin: shams@utm.my\n\nBritish Council Algeria:\nbritishcouncil.dz/ielts",
     "• Spain/Nordic: see Sheet EUROPE\n• Malaysia: see Sheet MALAYSIA\n• IELTS: book immediately\n• Presentation: prepare slides with your papers + SENTRY",
     "⬜ TODO"),
    ("JUNE 2026\nJun 9–30",
     """STEP 17 — Interview Phase:
→ Respond to ALL professor replies within 24 hours
→ Schedule video calls — prepare research presentation
→ Take IELTS exam (if booked for June)

STEP 18 — Formal Italy PhD applications:
→ Apply formally on dottorato.polito.it (PoliTo PNRR call expected June)
→ Apply on phd.polimi.it (PoliMi PNRR call)
→ Apply on dottorato.unibo.it (UniBO call)
→ Upload: CV + English transcripts + research proposal + 2 reference letters

STEP 19 — MEXT Japan submission:
→ Final MEXT application submitted to Japanese Embassy Algiers
→ Bring: completed forms + transcripts + 2 ref letters + research plan

STEP 20 — China CSC preparation:
→ Email Prof. Qiu at BUPT for pre-acceptance
→ CSC portal opens Feb 2027 — early contact needed now""",
     "All professors who replied\nPoliTo PhD portal: dottorato.polito.it\nPoliMi: phd.polimi.it\nBologna: dottorato.unibo.it\nJapanese Embassy Algiers",
     "• Formal applications: CV + transcripts (English ✅) + proposal + 2 ref letters\n• MEXT: completed forms package\n• No IELTS required for Italy/Japan applications",
     "⬜ TODO"),
    ("JULY–SEPT 2026\nMonthly",
     """STEP 21 — Send second wave of professor emails (any not yet contacted)
STEP 22 — Follow up all May/June emails (14-day rule)
STEP 23 — DAAD preparation (deadline Nov 1, 2026):
→ If German professor agreed: collect supervisor letter
→ Register at portal.daad.de
→ Write 5-page research proposal (SENTRY extended version)
STEP 24 — VLIR-UOS Belgium preparation (opens Nov 2026):
→ Build relationship with Prof. De Turck over email
→ Prepare VLIR documents package
STEP 25 — Interview season: most EU decisions come July–Sept""",
     "German professors for DAAD\nDe Turck for VLIR-UOS\nAll interview invitations",
     "DAAD package: supervisor letter + proposal + transcripts\nVLIR package: CV + proposal + 2 refs\nInterview: research presentation ready",
     "⬜ TODO"),
    ("OCT–NOV 2026\n🚨 DEADLINES",
     """CRITICAL DEADLINES THIS PERIOD:

November 1: DAAD application (HARD DEADLINE)
→ Submit at portal.daad.de
→ Must have German supervisor letter

November 15: VLIR-UOS Belgium opens (~Nov)
→ Go to vliruos.be — start application
→ Prof. De Turck must confirm supervision in VLIR system

Ongoing: Apply to any open MSCA/Horizon positions on EURAXESS
Ongoing: Canada Waterloo/Toronto (deadline Jan 15, 2027)""",
     "DAAD portal: portal.daad.de\nVLIR portal: scholarship.vliruos.be\nEUREXESS: check weekly",
     "DAAD: full package (see Sheet SCHOLARSHIPS)\nVLIR: full package\nAll English transcripts ✅ ready",
     "⬜ TODO"),
    ("DEC 2026 –\nFEB 2027",
     """STEP 26 — VLIR-UOS Belgium: complete and submit (deadline ~Mar 2027)
STEP 27 — Canada applications: Waterloo deadline Jan 15
STEP 28 — Korea GKS: embassy route Feb–Mar 2027
STEP 29 — China CSC portal opens: apply Feb–Apr 2027
STEP 30 — Singapore SINGA: Dec 2026 / Apr 2027 intake""",
     "Waterloo: grad.uwaterloo.ca\nKorean Embassy: 10 Rue Khelifa Boukhalfa, Algiers\nCSC: studyinchina.csc.edu.cn\nSINGA: a-star.edu.sg",
     "IELTS result needed for Canada/Korea/Singapore\nAll other documents ✅ ready",
     "⬜ TODO"),
    ("MAR–OCT 2027\nDecision + Enrollment",
     """STEP 31 — Evaluate all offers received (EU decisions Mar–May 2027)
STEP 32 — Accept best offer — notify all others within 7 days
STEP 33 — Begin visa process immediately after acceptance
STEP 34 — Submit IELTS certificate to chosen university (if conditional)
STEP 35 — PhD starts: Sep/Oct 2027""",
     "Chosen university admissions office",
     "Final IELTS certificate\nVisa application documents\nAcceptance letter + enrollment forms",
     "⬜ TODO"),
]

bgs = ["FDEDEC","FEF9E7","D6EAF8","EAFAF1","D5F5E3","F4ECF7","FFF3CD","EBF5FB","D5F5E3"]
for i, (when, action, who, what, status) in enumerate(plan):
    r = 3 + i
    ws1.row_dimensions[r].height = 100
    bg = bgs[i % len(bgs)]
    C(ws1, r, 1, when, bold=True, color="1A5276", bg=bg, sz=9, align="center")
    C(ws1, r, 2, action, bg=bg, sz=9)
    C(ws1, r, 3, who, color="117A65", bg=bg, sz=9)
    C(ws1, r, 4, what, color="6C3483", bg=bg, sz=9)
    C(ws1, r, 5, status, bg=bg, sz=9, align="center")

BD(ws1, 2, 2+len(plan), 1, 5)
ws1.freeze_panes = "A3"
W(ws1, [18, 56, 34, 38, 12])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ITALY (URGENT — PNRR opening NOW)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🇮🇹 ITALY (URGENT)")
ws2.sheet_view.showGridLines = False
T(ws2, "ITALY — PNRR PhD Calls Opening May–June 2026 | 5 Professors | Exact Emails | No IELTS Required", "C0392B", 5)

for i, h in enumerate(["PROFESSOR + LAB", "SEND TO (email)", "SUBJECT LINE", "COMPLETE EMAIL (copy → send)", "NOTES + DEADLINE"], 1):
    H(ws2, 2, i, h, bg="E74C3C", sz=9)
ws2.row_dimensions[2].height = 28

italy_emails = [
    ("Prof. Carla Fabiana Chiasserini\nPolitecnico di Torino\nNetAI Lab — DET\n\n⭐ MENTIONED IN YOUR SENTRY PROPOSAL\n'LLM-driven NOC automation'\n\nFunding: PNRR\n~€1,200/mo net",
     "chiasserini@polito.it",
     "PhD Inquiry 2026/27 — SENTRY: Runtime GNN-based Behavioural Monitor for Multi-Agent LLM Systems | 2 Scopus Papers | PNRR Call",
     """Dear Prof. Chiasserini,

I am writing to you directly because I mention your group's work on LLM-driven NOC automation in my PhD research proposal — and I believe SENTRY's objectives create a direct co-publication opportunity with the NetAI Lab.

My name is Abd El Madjid Kahoul — MSc graduate in Software Engineering and Intelligent Systems from the University of Constantine 2, Algeria (2025, Top Class B). My PhD research proposal is titled SENTRY: A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems in Operational Environments.

""" + SENTRY_BRIEF + """

The connection to your group is direct: SENTRY's stretch goal is validation of the monitoring framework in a Network Operations Centre (NOC) simulation environment — precisely the domain your group works on with LLM-driven NOC automation. A SENTRY-monitored multi-agent NOC pipeline would be a natural co-publication between our groups.

""" + PAPERS + """

I am applying for the Politecnico di Torino PNRR PhD call (opening June 2026 on dottorato.polito.it). Italy requires no IELTS at the application stage. """ + COND + """

Would you be open to discussing a potential PhD position in your group for the 2026/2027 PNRR call?

""" + SIG,
     "📅 Send: Week of May 12, 2026\n⏰ 9:00 AM CET Tuesday\n📎 Attach: CV (2 pages) + SENTRY_Proposal_v2_1.pdf\n🌐 Apply via: dottorato.polito.it (call opens ~June)\n💰 PNRR = ~€1,200/mo net, 3 years\n✅ NO IELTS at Italian PhD application\n🔁 Follow-up: +14 days"),

    ("Prof. Marco Montali\nFree University of Bozen-Bolzano\nFaculty of Computer Science\n\n🎯 Process Mining = Runtime Monitoring\n\nFunding: PNRR / unibz internal\n~€1,200/mo net",
     "marco.montali@unibz.it",
     "PhD Inquiry 2026/27 — SENTRY: Graph-Based Runtime Monitoring for Multi-Agent LLM Systems | Scopus Q2 | Process Mining Bridge",
     """Dear Prof. Montali,

Your work on process mining and runtime verification establishes the theoretical foundation that SENTRY — my PhD research proposal on runtime behavioural monitoring for multi-agent LLM systems — builds upon. I am writing to inquire whether you might be interested in supervising this direction.

I am Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The bridge to process mining is precise: SENTRY's agent interaction graph is essentially a runtime process model being compared against a learned normal-behaviour baseline — this is directly analogous to conformance checking in process mining. Where classical process mining operates on event logs of business processes, SENTRY applies the same paradigm to LLM agent execution traces. The graph representation of inter-agent communication (nodes = agents, edges = messages/tool calls) maps cleanly onto a process net with non-deterministic branching. I believe there is a compelling paper to be written comparing SENTRY's GNN-based anomaly detection with conformance-checking approaches on the same agent execution traces.

""" + PAPERS + """

Italy requires no IELTS at the PhD application stage. """ + COND + """

Would you be open to discussing a potential PhD position for 2026/2027?

""" + SIG,
     "📅 Send: Week of May 12, 2026\n⏰ 9:00 AM CET\n📎 Attach: CV + SENTRY proposal PDF\n💡 Key argument: process mining = conformance checking = SENTRY\n🌐 unibz.it/en/faculties/computer-science\n✅ NO IELTS required Italy\n🔁 Follow-up: +14 days"),

    ("Prof. Danilo Ardagna\nPolitecnico di Milano — DEIB\nCloud Lab\n\n🎯 Distributed AI + Cloud Systems\n\nFunding: PNRR\nphd.polimi.it (opening June)",
     "danilo.ardagna@polimi.it",
     "PhD Inquiry 2026/27 — SENTRY: Non-Invasive Runtime Monitoring for Distributed Multi-Agent LLM Systems | 2 Scopus Papers | PNRR",
     """Dear Prof. Ardagna,

I am writing because your research on AI for distributed systems and cloud infrastructure management creates a natural research environment for SENTRY — my PhD proposal on runtime behavioural reliability monitoring for multi-agent LLM systems deployed in operational settings.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The connection to your group is direct: multi-agent LLM systems are now deployed as distributed cloud workloads — on AWS Lambda chains, Azure AI pipelines, and LangGraph/AutoGen-based AIOps platforms. SENTRY's non-invasive monitoring framework (overhead < 5 ms per execution cycle, structural features only) is explicitly designed for production-scale deployment in these environments. Integrating SENTRY with your group's work on cloud resource management and AI-driven infrastructure would extend both lines of research.

""" + PAPERS + """

I am applying for the Politecnico di Milano PNRR PhD call (opening ~June 2026 at phd.polimi.it). Italy requires no IELTS at application stage. """ + COND + """

Would you have any PNRR-funded PhD openings in your group for 2026/2027?

""" + SIG,
     "📅 Send: Week of May 12, 2026\n⏰ 9:00 AM CET\n📎 Attach: CV + SENTRY proposal\n🌐 Apply: phd.polimi.it (June call)\n💰 PNRR = ~€1,200/mo, 3 years\n✅ NO IELTS Italy\n🔁 Follow-up: +14 days"),

    ("Prof. Federico Cabitza\nUniversità degli Studi di Milano-Bicocca\nDISCo Department\n\n🎯 XAI + Human-in-the-Loop AI\nTrustworthy AI expert\n\nFunding: PNRR / Italian PRIN",
     "federico.cabitza@unimib.it",
     "PhD Inquiry 2026/27 — SENTRY: XAI-Driven Operator-Readable Monitoring for Multi-Agent LLM Systems | 2 Scopus Papers | Trustworthy AI",
     """Dear Prof. Cabitza,

Your research on explainable AI, human-in-the-loop systems, and trustworthy AI evaluation directly addresses the explainability objectives of SENTRY — my PhD research proposal on runtime behavioural monitoring for multi-agent LLM systems.

I am Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The connection to your work is precise. SENTRY's explainability component is central, not peripheral: SHAP over node/edge features and GNNExplainer produce ranked lists of inter-agent interactions most responsible for each anomaly score. I propose a user study (n=12 professional software engineers) to evaluate whether these explanations satisfy the simulatability, decomposability, and algorithmic transparency criteria from the interpretable AI framework your group has helped establish. This user study design (n=12, pilot n=3, power analysis α=0.05 targeting d=0.5) is already structured in the SENTRY proposal. Your expertise in designing and evaluating XAI user studies would be invaluable.

""" + PAPERS + """

Italy requires no IELTS at the PhD application stage. """ + COND + """

Would you be open to discussing a potential PhD position for 2026/2027?

""" + SIG,
     "📅 Send: Week of May 19, 2026\n⏰ 9:00 AM CET\n📎 Attach: CV + SENTRY proposal\n💡 Key: user study design already in SENTRY — his speciality\n🌐 unimib.it/go/6EG6F9\n✅ NO IELTS Italy\n🔁 Follow-up: +14 days"),

    ("Prof. Nicola Tonellotto\nUniversità di Pisa — Dept CS\nHPC + Distributed Systems\n\n🎯 Inference efficiency for AI systems\nDistributed deployment\n\nFunding: PNRR / CNR Pisa",
     "nicola.tonellotto@unipi.it",
     "PhD Inquiry 2026/27 — SENTRY: Efficient Graph-Based Runtime Monitoring for Multi-Agent LLM Systems | 2 Scopus Papers",
     """Dear Prof. Tonellotto,

I am writing because your research on efficient inference for large AI models and distributed system performance directly addresses the engineering constraints of SENTRY — my PhD proposal on non-invasive runtime monitoring for multi-agent LLM systems.

I am Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The engineering constraint in SENTRY is non-trivial: the monitoring framework must operate with < 5 ms overhead per execution cycle, < 30 s inference latency, and false positive rate < 5% — otherwise it disrupts the very systems it monitors. Your expertise in model compression, efficient inference, and distributed deployment is directly relevant to making SENTRY's GAT-based detection model deployable at production scale. I see a clear research direction: deploying SENTRY as a lightweight sidecar process with quantized/pruned GNN models, co-designed with the inference scheduling of the monitored LLM agent chain.

""" + PAPERS + """

Italy requires no IELTS at the PhD application stage. """ + COND + """

Would you have any funded PhD openings for 2026/2027?

""" + SIG,
     "📅 Send: Week of May 19, 2026\n⏰ 9:00 AM CET\n📎 Attach: CV + SENTRY proposal\n🌐 unipi.it / CNR Pisa\n✅ NO IELTS Italy\n🔁 Follow-up: +14 days"),
]

for i, (prof, to, subj, body, notes) in enumerate(italy_emails):
    r = 3 + i
    ws2.row_dimensions[r].height = 320
    bg = ALT[i % len(ALT)]
    C(ws2, r, 1, prof, bold=True, color="C0392B", bg=bg, sz=9)
    C(ws2, r, 2, to, bold=True, color="117A65", bg=bg, sz=9)
    C(ws2, r, 3, subj, bold=True, color="6C3483", bg=bg, sz=9)
    C(ws2, r, 4, body, bg=bg, sz=9)
    C(ws2, r, 5, notes, italic=True, color="E67E22", bg=bg, sz=9)

# Italy PNRR How-To block
r_how = 3 + len(italy_emails) + 1
ws2.merge_cells(start_row=r_how, start_column=1, end_row=r_how, end_column=5)
x = ws2.cell(row=r_how, column=1, value="ITALY PNRR APPLICATION — EXACT STEPS")
x.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
x.fill = PatternFill("solid", fgColor="E74C3C")
x.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r_how].height = 24

pnrr_steps = """HOW TO APPLY FOR ITALY PNRR PhD (exact process — use your English transcripts):

STEP 1 — Right now: email the professor (use email bodies above). Goal: get "I'm interested in supervising you" reply.
STEP 2 — May–June 2026: Check these portals DAILY for PNRR calls:
   → Politecnico di Torino: dottorato.polito.it
   → Politecnico di Milano: phd.polimi.it
   → University of Bologna: dottorato.unibo.it
   → Free University Bolzano: phd.unibz.it
   → University of Pisa: phd.unipi.it
STEP 3 — When call opens: register on the portal + download the "bando di dottorato" (call document).
STEP 4 — Find the position in the list that matches your research area (look for: AI, ML, trustworthy AI, distributed systems, network intelligence).
STEP 5 — Upload documents (ALL already have English translations ✅):
   → Research proposal: SENTRY PDF (v2.1) — already written!
   → CV: 2-page academic version
   → MSc transcript in English ✅
   → BSc transcript in English ✅
   → MSc diploma (certified copy)
   → 2 reference letters (MSc supervisor + Ooredoo Algeria manager)
   → Passport copy
   → NO IELTS required — language is not required at application stage in Italy
STEP 6 — Selection: Italian PhD selection is usually a thesis proposal defense (15 min in English). Prepare your SENTRY presentation.
STEP 7 — Results: 4–6 weeks after call closes. If awarded: PNRR contract ~€1,200/mo net for 3 years."""

ws2.merge_cells(start_row=r_how+1, start_column=1, end_row=r_how+1, end_column=5)
y = ws2.cell(row=r_how+1, column=1, value=pnrr_steps)
y.font = Font(name="Arial", size=9)
y.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
y.fill = PatternFill("solid", fgColor="FADBD8")
ws2.row_dimensions[r_how+1].height = 160

BD(ws2, 2, r_how+1, 1, 5)
ws2.freeze_panes = "A3"
W(ws2, [28, 26, 36, 92, 30])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — JAPAN (URGENT — MEXT embassy route open NOW)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🇯🇵 JAPAN (MEXT OPEN NOW)")
ws3.sheet_view.showGridLines = False
T(ws3, "JAPAN — MEXT Scholarship Embassy Route OPEN NOW May–June 2026 | 5 Professors | No IELTS", "1A3A5C", 5)

for i, h in enumerate(["PROFESSOR + LAB", "SEND TO (email)", "SUBJECT LINE", "COMPLETE EMAIL (copy → send)", "NOTES + MEXT DEADLINE"], 1):
    H(ws3, 2, i, h, bg="2C3E7A", sz=9)
ws3.row_dimensions[2].height = 28

japan_emails = [
    ("Prof. Yutaka Matsuo\nUniversity of Tokyo\nWeb Intelligence Lab (Matsuo Lab)\n\n🎯 #1 LLM researcher in Japan\nAI agent reliability + LLM systems\n\nFunding: MEXT + JSPS + industry",
     "matsuo@weblab.t.u-tokyo.ac.jp",
     "PhD Inquiry 2026/27 — SENTRY: Runtime Reliability Monitor for Multi-Agent LLM Systems | Scopus Q2 | MEXT Scholarship Applicant",
     """Dear Prof. Matsuo,

You are Japan's leading researcher in large language models and agent AI systems. SENTRY — my PhD research proposal — addresses what I believe is the most critical unsolved reliability problem your field is now encountering at the deployment stage.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The problem SENTRY addresses has emerged directly from the deployment of the research your lab has pioneered. AutoGen-based and LangGraph-based multi-agent systems have moved from academic benchmarks into commercial AIOps, NOC, and software development pipelines — yet the empirical work of Cemri et al. (ICLR 2025 Workshop) documents 14 distinct failure modes with task success rates as low as 25%, and these failures are structural (not correctable by prompt engineering). SENTRY is the first non-invasive framework designed specifically for this reliability gap.

""" + PAPERS + """

I plan to apply for the MEXT Scholarship through the university recommendation route (Todai). I am writing to inquire whether you would be open to considering me for your lab, and whether you would be willing to provide a MEXT university nomination.

Japan does not require IELTS for MEXT. My two Scopus publications demonstrate English research proficiency. """ + COND + """

Would you be open to a brief conversation about potential PhD opportunities?

""" + SIG,
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM JST (Japan Standard Time = CET +7h)\n📎 Attach: CV + SENTRY proposal PDF\n⚠️ MEXT university route: Oct–Nov 2026\n⚠️ MEXT EMBASSY route: contact embassy NOW!\n💰 MEXT: ¥145,000/mo + tuition + airfare\n✅ NO IELTS for MEXT\n🔁 Follow-up: +14 days"),

    ("Prof. Masashi Sugiyama\nRIKEN Center for AI Project (AIP)\nUniversity of Tokyo (joint appointment)\n\n🎯 Anomaly detection + ML theory\nDistribution shift, out-of-distribution\n\nFunding: JSPS + RIKEN + MEXT",
     "sugi@k.u-tokyo.ac.jp",
     "PhD Inquiry 2026/27 — SENTRY: GNN-Based Behavioural Anomaly Detection for Multi-Agent LLM Systems | Scopus Q2 | MEXT Applicant",
     """Dear Prof. Sugiyama,

Your foundational research on anomaly detection, distribution shift, and out-of-distribution detection addresses the exact theoretical challenge that SENTRY — my PhD proposal on runtime monitoring for multi-agent LLM systems — must solve in a novel graph-structured domain.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The theoretical challenge SENTRY poses is this: normal agent behaviour is characterised by a learned distribution over interaction graph structures. Anomalous executions correspond to out-of-distribution graph events — a reasoning loop creates anomalously dense cycles, a hallucinated tool call produces an edge type inconsistent with the learned structure, delegation drift shifts the graph's degree distribution. This is formally an out-of-distribution detection problem on dynamic graph-structured data — precisely the intersection of your work on distribution shift and the temporal graph anomaly detection literature (TGN, Anomal-E). The reconstruction-based approach in SENTRY (autoencoder on graph snapshots; anomaly = high reconstruction error) has a direct theoretical connection to your work on deep anomaly detection.

""" + PAPERS + """

I intend to apply for the MEXT Scholarship. MEXT requires no IELTS. My publications demonstrate English research proficiency. """ + COND + """

Would you be open to discussing a potential PhD position in your group?

""" + SIG,
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM JST\n📎 Attach: CV + SENTRY proposal\n💡 RIKEN AIP = top AI research institute in Japan\n💰 MEXT: ¥145,000/mo + full tuition\n✅ NO IELTS\n🔁 Follow-up: +14 days"),

    ("Prof. Satoshi Nakamura\nNara Institute of Science and Technology (NAIST)\nAI & Human Lab / Augmented Human Comm. Lab\n\n🎯 Multi-agent AI, human-AI interaction\nAgent system reliability\n\nFunding: JSPS + MEXT",
     "s-nakamura@is.naist.jp",
     "PhD Inquiry 2026/27 — SENTRY: Runtime Monitoring for Multi-Agent AI Systems | Scopus Q2 | MEXT Scholarship Applicant",
     """Dear Prof. Nakamura,

Your research on multi-agent AI systems and human-AI collaboration is directly relevant to SENTRY — my PhD proposal on runtime behavioural monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The human dimension of SENTRY is significant: the HANDOFF intervention level escalates detected anomalies to human operators with an XAI-generated explanation report. This is a human-in-the-loop design where the monitoring system must (a) detect the failure, (b) explain which agent interactions caused it, and (c) present this to a human in a form that enables rapid, correct decision-making. The operator utility study (n=12, measuring simulatability, decomposability, algorithmic transparency) is a direct contribution to human-AI collaboration research.

""" + PAPERS + """

I plan to apply for the MEXT Scholarship (university recommendation route at NAIST). MEXT requires no IELTS. My publications demonstrate English research proficiency. """ + COND + """

Would you be open to a conversation about potential PhD opportunities?

""" + SIG,
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM JST\n📎 Attach: CV + SENTRY proposal\n🌐 NAIST: naist.ac.jp/en\n💰 MEXT: ¥145,000/mo + tuition\n✅ NO IELTS\n🔁 Follow-up: +14 days"),

    ("Prof. Katsumi Tanaka\nKyoto University\nGraduate School of Informatics\n\n🎯 Knowledge graphs, AI systems\nAutonomous agents, semantic web\n\nFunding: JSPS + MEXT",
     "tanaka.katsumi.7r@kyoto-u.ac.jp",
     "PhD Inquiry 2026/27 — SENTRY: Knowledge-Graph-Enhanced Monitoring for Multi-Agent LLM Systems | Scopus Q2 | MEXT",
     """Dear Prof. Tanaka,

Your research on knowledge representation, semantic systems, and autonomous agent architectures intersects with SENTRY — my PhD proposal on runtime monitoring for multi-agent LLM systems — in a direction I find scientifically compelling.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

An extension I envision for SENTRY at Kyoto: enriching the agent interaction graph with semantic knowledge — tool taxonomies, task ontologies, expected delegation chains — as prior knowledge for the GAT anomaly detector. Where SENTRY's current design learns normality purely from structural graph statistics, a knowledge-enriched version would combine structural anomaly scores with semantic consistency checks ("was this tool call semantically consistent with the agent's declared objective?"). This Knowledge Graph-augmented SENTRY (KG-SENTRY) would be a natural direction given your group's expertise.

""" + PAPERS + """

I intend to apply for MEXT (university recommendation or embassy route). MEXT requires no IELTS. """ + COND + """

Would you have any PhD openings for 2026/2027?

""" + SIG,
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM JST\n📎 Attach: CV + SENTRY proposal\n🌐 i.kyoto-u.ac.jp\n💰 MEXT: ¥145,000/mo\n✅ NO IELTS\n🔁 Follow-up: +14 days"),

    ("Prof. Takafumi Kanamori\nTokyo Institute of Technology (Tokyo Tech)\nDept of Mathematical and Computing Science\n\n🎯 Anomaly detection, density ratio estimation\nStatistical ML, change-point detection\n\nFunding: JSPS + MEXT",
     "kanamori@c.titech.ac.jp",
     "PhD Inquiry 2026/27 — SENTRY: Statistical Anomaly Detection in Agent Interaction Graphs | Scopus Q2 | MEXT Applicant",
     """Dear Prof. Kanamori,

Your research on statistical anomaly detection, density ratio estimation, and change-point detection provides the probabilistic foundations for a key open question in SENTRY — my PhD proposal on runtime behavioural monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

A specific open problem in SENTRY where your expertise is directly relevant: threshold calibration for the intervention engine. SENTRY's current design proposes ROC-based operating-point selection and Bayesian threshold estimation for the Monitor/Pause/Rollback confidence levels — but the theoretical properties of these thresholds under concept drift (as LLM model weights update or agent task distributions shift) are unknown. Extending your work on density ratio estimation and non-stationary distribution modelling to the agent interaction graph domain would be a principled theoretical contribution with direct practical impact.

""" + PAPERS + """

I intend to apply for MEXT (university recommendation route at Tokyo Tech). MEXT requires no IELTS. """ + COND + """

Would you be open to discussing PhD opportunities in your group?

""" + SIG,
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM JST\n📎 Attach: CV + SENTRY proposal\n🌐 titech.ac.jp/en\n💰 MEXT: ¥145,000/mo + tuition + airfare\n✅ NO IELTS\n🔁 Follow-up: +14 days"),
]

for i, (prof, to, subj, body, notes) in enumerate(japan_emails):
    r = 3 + i
    ws3.row_dimensions[r].height = 320
    bg = ALT[i % len(ALT)]
    C(ws3, r, 1, prof, bold=True, color="1A3A5C", bg=bg, sz=9)
    C(ws3, r, 2, to, bold=True, color="117A65", bg=bg, sz=9)
    C(ws3, r, 3, subj, bold=True, color="6C3483", bg=bg, sz=9)
    C(ws3, r, 4, body, bg=bg, sz=9)
    C(ws3, r, 5, notes, italic=True, color="E67E22", bg=bg, sz=9)

# Japan MEXT how-to block
rj = 3 + len(japan_emails) + 1
ws3.merge_cells(start_row=rj, start_column=1, end_row=rj, end_column=5)
jh = ws3.cell(row=rj, column=1, value="MEXT SCHOLARSHIP — EXACT STEPS (Embassy Route — APPLY THIS WEEK)")
jh.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
jh.fill = PatternFill("solid", fgColor="2C3E7A")
jh.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[rj].height = 24

mext_steps = """HOW TO APPLY FOR MEXT SCHOLARSHIP — EMBASSY ROUTE (deadline: May–June 2026 — URGENT)

STEP 1 — THIS WEEK: Go to Japanese Embassy in Algiers
→ Address: 25 Chemin Abdelkader Gadouche, Hydra, Algiers (near the Hydra embassies district)
→ Phone: +213 21 54 04 08
→ Email: jp-emb-algiers@ws.mofa.go.jp
→ Say: "I want to apply for the 2026 MEXT Research Student Scholarship (Embassy Recommendation)"
→ Request: application forms + exact current deadline for Algeria (deadline changes each year)

STEP 2 — Download MEXT Application Form:
→ mext.go.jp/en/policy/education/highered/title02/detail02/sdetail02/1373897.htm
→ Fill: personal info, intended university (pick from Japan emails above), supervisor name, research area

STEP 3 — Write Research Plan (2 pages, in English — the most important document):
→ Title: "SENTRY: Runtime Behavioural Reliability Monitoring for Multi-Agent LLM Systems"
→ Section 1: Your background (MSc + 2 Scopus publications)
→ Section 2: Research problem (silent failures in multi-agent LLM systems)
→ Section 3: Proposed methodology (GAT + GNNExplainer + intervention engine)
→ Section 4: Why Japan / why this professor / why MEXT
→ Use your SENTRY_Proposal_v2_1.pdf as the source — summarize it in 2 pages

STEP 4 — Gather documents (you ALREADY HAVE all of these with English translations ✅):
→ ✅ MSc transcript in English
→ ✅ BSc transcript in English
→ MSc and BSc diplomas (certified copies)
→ 2 recommendation letters (MSc supervisor + Ooredoo Algeria manager) — ask NOW
→ MEXT health form (download from embassy, fill at hospital in Constantine or Algiers)
→ Passport copy (valid 2+ years)
→ Passport-size photos
→ NO IELTS required for MEXT — ever

STEP 5 — Submit to Japanese Embassy Algiers by their stated deadline (typically May–June)

STEP 6 — Embassy screening: language test + interview at embassy (in English)

STEP 7 — If selected by embassy: your application goes to MEXT ministry + then to university
STEP 8 — You start as "research student" (non-degree) at the Japanese university → convert to PhD after 1 semester

FUNDING: ¥145,000/mo (~€900) + full tuition waiver + round-trip airfare + settling-in allowance"""

ws3.merge_cells(start_row=rj+1, start_column=1, end_row=rj+1, end_column=5)
jb = ws3.cell(row=rj+1, column=1, value=mext_steps)
jb.font = Font(name="Arial", size=9)
jb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
jb.fill = PatternFill("solid", fgColor="D6EAF8")
ws3.row_dimensions[rj+1].height = 190

BD(ws3, 2, rj+1, 1, 5)
ws3.freeze_panes = "A3"
W(ws3, [28, 26, 36, 92, 30])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — MALAYSIA (MIS — check if open NOW)
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🇲🇾 MALAYSIA (MIS CHECK)")
ws4.sheet_view.showGridLines = False
T(ws4, "MALAYSIA — MIS Scholarship (Check NOW if Open) | 5 Professors | Multi-Agent AI + Trustworthy AI", "1A7A3A", 5)

for i, h in enumerate(["PROFESSOR + LAB", "SEND TO (email)", "SUBJECT LINE", "COMPLETE EMAIL (copy → send)", "NOTES + MIS DEADLINE"], 1):
    H(ws4, 2, i, h, bg="27AE60", sz=9)
ws4.row_dimensions[2].height = 28

malaysia_emails = [
    ("Prof. Mohd Sharifuddin Ahmad\nUniversiti Teknologi Malaysia (UTM)\nFaculty of Engineering — CSISC\n\n🎯 Multi-agent systems research group\nIntelligent agents, MAS coordination\n\nFunding: MIS + UTM research grants",
     "shams@utm.my",
     "PhD Inquiry 2026/27 — SENTRY: Behavioural Reliability Monitoring for Multi-Agent LLM Systems | Scopus Q2 | MIS Scholarship Applicant",
     """Dear Prof. Ahmad,

Your research on multi-agent systems and intelligent agent coordination at UTM's CSISC group is directly relevant to SENTRY — my PhD research proposal on runtime behavioural reliability monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The connection to your research area is foundational: SENTRY addresses the reliability and coordination failure modes in deployed multi-agent LLM systems — the exact failure patterns (reasoning loops = inter-agent misalignment, tool-call anomalies = task verification failure, delegation drift = system design failure) that your group studies in classical MAS settings, now emerging at scale in LLM-based deployments. SENTRY's graph representation of agent interactions maps directly onto the coordination protocol graphs studied in multi-agent systems research.

""" + PAPERS + """

I plan to apply for the Malaysia International Scholarship (MIS) and would need a supervisor willing to support my application. MIS does not require IELTS at the application stage. """ + COND + """

Would you be open to discussing a PhD position and MIS supervision for 2026/2027?

""" + SIG,
     "📅 Send: Week of June 2, 2026\n⏰ Morning email\n📎 Attach: CV + SENTRY proposal\n⚠️ CHECK NOW: scholarship.moe.gov.my\n💰 MIS = Full tuition + RM 3,000/mo + health + airfare\n✅ NO IELTS at MIS application\n🔁 Follow-up: +14 days"),

    ("Prof. Shahrul Azman Mohd Noah\nUniversiti Kebangsaan Malaysia (UKM)\nFaculty of Information Science and Technology\n\n🎯 AI systems, NLP, intelligent knowledge systems\nSemantic AI, ontology-based reasoning\n\nFunding: MIS + UKM research",
     "shahrul@ukm.edu.my",
     "PhD Inquiry 2026/27 — SENTRY: Knowledge-Aware Runtime Monitoring for Multi-Agent LLM Systems | Scopus Q2 | MIS Applicant",
     """Dear Prof. Shahrul,

Your research on intelligent knowledge systems and semantic AI at UKM provides the knowledge representation foundations for an important extension of SENTRY — my PhD proposal on runtime monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

SENTRY currently detects anomalies from structural graph statistics only. An important open direction is incorporating semantic knowledge — what tool calls were expected given the agent's declared objective? What delegation patterns are semantically consistent with the task decomposition plan? Enriching SENTRY's detection model with ontology-based constraints would reduce false positives and improve explanation quality. Your expertise in knowledge representation and semantic systems would be essential for this direction.

""" + PAPERS + """

I plan to apply for the Malaysia International Scholarship (MIS) with UKM as the host institution. MIS does not require IELTS at the application stage. """ + COND + """

Would you be open to discussing PhD opportunities and MIS supervision?

""" + SIG,
     "📅 Send: Week of June 2, 2026\n⏰ Morning email\n📎 Attach: CV + SENTRY proposal\n💡 Propose semantic extension of SENTRY — his speciality\n💰 MIS = Full tuition + RM 3,000/mo\n✅ NO IELTS for MIS\n🔁 Follow-up: +14 days"),

    ("Prof. Kasturi Dewi Varathan\nUniversiti Malaya (UM)\nFaculty of Computer Science and IT\n\n🎯 Trustworthy AI, explainable AI\nData-driven decision systems\n\nFunding: MIS + UM UMRG",
     "kasturi@um.edu.my",
     "PhD Inquiry 2026/27 — SENTRY: XAI-Based Trustworthy Monitoring for Multi-Agent LLM Systems | Scopus Q2 | MIS Scholarship",
     """Dear Prof. Varathan,

Your research on trustworthy AI and explainable decision-making systems directly addresses the explainability design objectives of SENTRY — my PhD proposal on runtime behavioural monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

SENTRY's explainability component is central to its trustworthiness claims: SHAP over node/edge features and GNNExplainer/PGExplainer attribution on agent interaction graphs. The operator utility study (n=12 professional software engineers, measuring simulatability, decomposability, and algorithmic transparency) is a direct trustworthy AI evaluation. Your expertise in designing human-centred explainability evaluations would be invaluable for validating SENTRY's explanation quality.

""" + PAPERS + """

I plan to apply for the Malaysia International Scholarship (MIS) with UM as host. MIS does not require IELTS. """ + COND + """

Would you have PhD openings for 2026/2027?

""" + SIG,
     "📅 Send: Week of June 2, 2026\n📎 Attach: CV + SENTRY proposal\n💡 Connect to her trustworthy AI expertise\n💰 MIS: full funding\n✅ NO IELTS for MIS\n🔁 Follow-up: +14 days"),

    ("Prof. Roliana Ibrahim\nUniversiti Teknologi Malaysia (UTM)\nFaculty of Engineering — Software Engineering\n\n🎯 Software engineering, AI systems\nIntelligent software, agent-based systems\n\nFunding: MIS + UTM grants",
     "roliana@utm.my",
     "PhD Inquiry 2026/27 — SENTRY: Non-Invasive Runtime Monitoring for Agent-Based Software Systems | Scopus Q2 | MIS",
     """Dear Prof. Ibrahim,

Your research on intelligent software engineering and agent-based systems positions SENTRY — my PhD proposal on non-invasive runtime monitoring for multi-agent LLM systems — as a natural extension of software engineering reliability principles to the LLM agent deployment context.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

From a software engineering perspective, SENTRY addresses the observability gap in agentic software: existing tools (LangSmith, Datadog LLM Observability, Langfuse) log what happened, but SENTRY answers whether the execution was anomalous relative to learned normal behaviour. This is the missing piece between tracing tools (software observability) and automated remediation (self-healing software). SENTRY's non-invasive callback-based instrumentation architecture — no modification to monitored agent code, < 5 ms overhead — follows established software engineering principles for monitoring system design.

""" + PAPERS + """

I plan to apply for the MIS with UTM as host. MIS does not require IELTS. """ + COND + """

Would you be open to discussing PhD opportunities and MIS supervision?

""" + SIG,
     "📅 Send: Week of June 2, 2026\n📎 Attach: CV + SENTRY proposal\n💡 Frame SENTRY as a software engineering contribution\n💰 MIS: full funding RM 3,000/mo\n✅ NO IELTS\n🔁 Follow-up: +14 days"),

    ("Prof. Noorminshah A. Iahad\nUniversiti Teknologi Malaysia (UTM)\nFaculty of Engineering — IT Management\n\n🎯 AI adoption, human-computer interaction\nIT governance, intelligent systems\n\nFunding: MIS + UTM",
     "noorminshah@utm.my",
     "PhD Inquiry 2026/27 — SENTRY: Human-in-the-Loop Reliability Governance for Multi-Agent LLM Systems | Scopus Q2 | MIS",
     """Dear Prof. Iahad,

Your research on human-computer interaction, IT governance, and AI adoption directly connects to SENTRY's human-in-the-loop design — my PhD proposal on runtime monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

SENTRY's HANDOFF intervention level — which escalates detected anomalies to human operators with an XAI-generated explanation report — is an IT governance mechanism: it defines when AI systems should defer to human judgment, how that deferral is communicated, and how human decisions feed back into the monitoring system. This connects to your work on IT governance and responsible AI adoption. Additionally, SENTRY's operator utility study (measuring whether practitioners can interpret GNN-based explanations) is a human-centred evaluation of an AI reliability tool — directly relevant to your HCI expertise.

""" + PAPERS + """

I plan to apply for the MIS. MIS does not require IELTS. """ + COND + """

Would you be open to discussing a potential PhD position?

""" + SIG,
     "📅 Send: Week of June 2, 2026\n📎 Attach: CV + SENTRY proposal\n💡 Frame SENTRY as HCI + governance\n💰 MIS: full tuition + RM 3,000/mo + health + airfare\n✅ NO IELTS for MIS\n🔁 Follow-up: +14 days"),
]

for i, (prof, to, subj, body, notes) in enumerate(malaysia_emails):
    r = 3 + i
    ws4.row_dimensions[r].height = 300
    bg = ALT[i % len(ALT)]
    C(ws4, r, 1, prof, bold=True, color="1A7A3A", bg=bg, sz=9)
    C(ws4, r, 2, to, bold=True, color="117A65", bg=bg, sz=9)
    C(ws4, r, 3, subj, bold=True, color="6C3483", bg=bg, sz=9)
    C(ws4, r, 4, body, bg=bg, sz=9)
    C(ws4, r, 5, notes, italic=True, color="E67E22", bg=bg, sz=9)

rm = 3 + len(malaysia_emails) + 1
ws4.merge_cells(start_row=rm, start_column=1, end_row=rm, end_column=5)
mh = ws4.cell(row=rm, column=1, value="MALAYSIA MIS SCHOLARSHIP — EXACT STEPS (CHECK IF OPEN NOW)")
mh.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
mh.fill = PatternFill("solid", fgColor="27AE60")
mh.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[rm].height = 24

mis_steps = """MALAYSIA INTERNATIONAL SCHOLARSHIP (MIS) — EXACT APPLICATION PROCESS

STEP 1 — RIGHT NOW: Go to scholarship.moe.gov.my
→ Check if the 2026 MIS cycle is currently open (typical window: March–May/June each year)
→ If open: you must apply THIS WEEK
→ If closed for 2026: next cycle opens March 2027 — still email professors now to prepare

STEP 2 — Email Malaysian professors (use bodies above) SIMULTANEOUSLY with MIS check:
→ Goal: get a supervisor willing to support your MIS application
→ UTM has a strong multi-agent systems group (Prof. Ahmad)
→ UM has trustworthy AI (Prof. Varathan)

STEP 3 — When professor agrees, fill MIS online application at scholarship.moe.gov.my:
→ Proposed university: UTM (Universiti Teknologi Malaysia) or UM (Universiti Malaya) or UKM
→ Proposed supervisor: name and email of professor who agreed
→ Research area: Trustworthy AI / Multi-Agent Systems / AI Reliability

STEP 4 — Upload documents (all ready with English translations ✅):
→ ✅ MIS online application form (completed at portal)
→ ✅ MSc transcript in English
→ ✅ BSc transcript in English
→ MSc and BSc diplomas (certified copies)
→ CV (2 pages, academic)
→ Personal statement / research proposal: adapt SENTRY proposal summary (2–3 pages)
   → Frame for Malaysia: "AI Reliability for Southeast Asian Digital Infrastructure"
→ 2 recommendation letters (MSc supervisor + Ooredoo manager) — on official letterhead, signed
→ Passport copy (color)
→ Health certificate (if required)
→ NO IELTS required for MIS application

STEP 5 — MIS evaluation: interview with selection committee
STEP 6 — If awarded: you receive full scholarship confirmation from Ministry of Education Malaysia
STEP 7 — Apply for Malaysian student visa (eVisa system)

FUNDING COVERAGE: Full tuition + RM 3,000/mo (~€600) + health insurance + round-trip airfare + settling-in allowance
DURATION: Up to 4 years (for PhD)
ELIGIBLE NATIONALS: Algeria IS eligible for MIS"""

ws4.merge_cells(start_row=rm+1, start_column=1, end_row=rm+1, end_column=5)
mb = ws4.cell(row=rm+1, column=1, value=mis_steps)
mb.font = Font(name="Arial", size=9)
mb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
mb.fill = PatternFill("solid", fgColor="D5F5E3")
ws4.row_dimensions[rm+1].height = 180

BD(ws4, 2, rm+1, 1, 5)
ws4.freeze_panes = "A3"
W(ws4, [28, 26, 36, 92, 30])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — EUROPE (Belgium, France, Germany, Netherlands, Austria, Sweden)
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🇪🇺 EUROPE")
ws5.sheet_view.showGridLines = False
T(ws5, "EUROPE — Belgium · France · Germany · Netherlands · Austria · Sweden | Precise SENTRY Emails | Linked Scholarships", "1F4E79", 6)

for i, h in enumerate(["COUNTRY + PROFESSOR + LAB", "EMAIL", "SUBJECT LINE", "COMPLETE EMAIL (copy → send)", "SCHOLARSHIP LINKED", "NOTES"], 1):
    H(ws5, 2, i, h, bg="2471A3", sz=9)
ws5.row_dimensions[2].height = 28

europe_emails = [
    ("🇧🇪 Belgium\nProf. Filip De Turck\nGhent University — IDLab\n\n⭐ EXPLICITLY NAMED in SENTRY proposal\n'IDLab has established record in distributed\napplication management and runtime\nobservability'\n\n🔗 Scholarship: VLIR-UOS",
     "Filip.DeTurck@ugent.be",
     "PhD Inquiry 2026/27 — SENTRY: Runtime GNN Monitor for Multi-Agent LLM Systems | IDLab Named in Proposal | VLIR-UOS Candidate",
     """Dear Prof. De Turck,

I am writing to you because IDLab is specifically named in my PhD research proposal as the primary target laboratory for SENTRY. Allow me to explain why.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B). My PhD research proposal is titled SENTRY: A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems in Operational Environments.

""" + SENTRY_BRIEF + """

In my proposal, I wrote the following about IDLab: "IDLab has an established record in distributed application management and runtime observability and is a leader in AI-driven network management. SENTRY's monitoring architecture directly extends their work on autonomous system observability to the LLM agent execution layer." This assessment comes from reading your group's work, particularly Sebrechts et al. on orchestrator conversation distributed management and your ongoing EU Horizon work on zero-touch network management. SENTRY's callback-based instrumentation architecture and graph-based anomaly detection model translate your observability framework from network telemetry to LLM agent execution traces.

The VLIR-UOS scholarship explicitly supports researchers from Algeria — I am a strong candidate and plan to apply in the November 2026 window with Ghent University as my nominated institution.

""" + PAPERS + """

""" + COND + """

Would you be open to a conversation about PhD opportunities in IDLab for 2026/2027?

""" + SIG,
     "VLIR-UOS Scholarship\n(Belgium)\nOpens: ~Nov 2026\nDeadline: ~Mar 31, 2027\nPortal: vliruos.be\nFull funding + €900/mo",
     "📅 Send: Week of May 19, 2026\n⏰ 8:00 AM CET Tuesday\n📎 CV + SENTRY proposal PDF\n🔁 Follow-up: +14 days\n💡 CC: Chris.Develder@ugent.be"),

    ("🇦🇹 Austria\nProf. Schahram Dustdar\nTU Wien — DSG Lab\n\n⭐ EXPLICITLY NAMED in SENTRY proposal\n'Dustdar's group works on edge AI orchestration\nand autonomous distributed systems with\nactive EU Horizon projects on trustworthy AI'\n\n🔗 Scholarship: FWF/FFG/EU Horizon",
     "dustdar@dsg.tuwien.ac.at",
     "PhD Inquiry 2026/27 — SENTRY: EU-AI-Act-Compliant Runtime Monitor for Multi-Agent LLM Systems | DSG Named in Proposal",
     """Dear Prof. Dustdar,

I am writing to you because the DSG Lab is specifically named in my PhD research proposal as a target laboratory for SENTRY. Let me explain the precise connection.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

In my proposal, I noted: "Dustdar's group works on edge AI orchestration and autonomous distributed systems with active EU Horizon projects on trustworthy AI infrastructure. SENTRY's compliance-oriented audit trail and intervention design align with their EU AI Act implementation research." This was based on reading your group's work, including Lackinger et al. on inference-aware orchestration for hierarchical federated learning. SENTRY's HANDOFF intervention engine and immutable audit trail are explicitly designed to comply with EU AI Act Article 12 (transparency and logging requirements for high-risk AI systems). Your group's expertise in EU AI Act implementation and trustworthy AI infrastructure would be essential for the compliance dimension of SENTRY.

""" + PAPERS + """

""" + COND + """

Would you be open to discussing a potential PhD position in the DSG Lab?

""" + SIG,
     "FWF (Austrian Science Fund)\nFFG (Austrian Research Promotion)\nEU Horizon Europe\n\nNo specific scholarship\n— direct lab funding\nPhD salary: ~€2,000/mo",
     "📅 Send: Week of May 19, 2026\n⏰ 9:00 AM CET\n📎 CV + SENTRY proposal PDF\n💡 Austria = no IELTS at application\n🔁 Follow-up: +14 days"),

    ("🇫🇷 France\nProf. Jérôme François\nInria Nancy — LORIA\nTeam RESIST\n\n🎯 Network anomaly detection + XAI\nBRIDGE: your papers → SENTRY\n\n🔗 Funding: Inria PhD contract\n~€2,100/mo (employee status!)",
     "jerome.francois@inria.fr",
     "[FR] Candidature Doctorat Inria RESIST — SENTRY : Monitoring comportemental par GNN pour systèmes multi-agents LLM | 2 publications Scopus",
     """Cher Professeur François,

Je vous contacte car mes deux publications Scopus en détection d'anomalies réseau — directement liées aux thèmes de RESIST — constituent le socle méthodologique d'un nouveau projet de recherche doctorale (SENTRY) qui étend cette approche à un nouveau domaine critique : la surveillance comportementale des systèmes multi-agents à base de LLM.

Je suis Abd El Madjid Kahoul — Master en Génie Logiciel et Systèmes Intelligents, Université de Constantine 2, Algérie (2025, mention bien).

PONT DIRECT ENTRE MES TRAVAUX ET SENTRY :

""" + SENTRY_BRIEF + """

Le fil conducteur est précis : mes deux publications Scopus (détection d'anomalies par DL hybride dans les réseaux LTE, et plateforme XAI RAVA avec SHAP) ont développé mes compétences en détection d'anomalies sur données opérationnelles réelles et en explicabilité des décisions IA. SENTRY applique ces mêmes compétences à un nouveau domaine — les graphes d'interaction inter-agents LLM — en remplaçant les flux KPI multivariés par des graphes dynamiques d'exécution, et en substituant le détecteur DL hybride par un réseau d'attention sur graphes (GAT). La logique méthodologique est identique ; le domaine applicatif change.

""" + PAPERS + """

J'ai le certificat TCF français niveau B2 et je suis en mesure de conduire mes travaux en français. Je prépare l'IELTS et puis fournir l'attestation avant l'inscription (acceptation conditionnelle).

Seriez-vous disponible pour un échange sur d'éventuelles opportunités doctorales à RESIST pour 2026/2027 ?

Avec mes respectueuses salutations,
Abd El Madjid Kahoul
abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17
DOI 1 : 10.18080/jtde.v13n4.1363 | DOI 2 : 10.18080/jtde.v14n1.142""",
     "Inria PhD Contract\n(Not scholarship — salary!)\n~€2,100/mo net\n\nNo scholarship app needed\n— apply directly at\njobs.inria.fr\n\nRolling applications",
     "📅 Envoyer: Mai 2026 mardi matin\n🇫🇷 ÉCRIRE EN FRANÇAIS — TCF B2!\n💡 CC: abdelkader.lahmadi@loria.fr\n📎 CV + SENTRY proposal\n🔁 Relance: +14 jours\n💰 Inria = contrat salarié ~€2,100/mo"),

    ("🇩🇪 Germany\nProf. Klaus-Robert Müller\nTU Berlin — Machine Learning Group\n\n⭐ SHAP co-inventor\nCited in your SENTRY proposal!\nXAI pioneer — directly relevant\n\n🔗 Scholarship: DAAD (deadline Nov 1, 2026)",
     "klaus-robert.mueller@tu-berlin.de",
     "PhD Inquiry 2026/27 — SENTRY: GNNExplainer vs SHAP vs PGExplainer for Agent Interaction Graph Explainability | Scopus Q2 | DAAD",
     """Dear Prof. Müller,

SHAP — the unified framework for feature attribution you co-developed — is cited in my PhD research proposal SENTRY as one of three explainability methods to be compared empirically in a new and previously unstudied domain: multi-agent LLM execution graphs.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The specific contribution relevant to your group: SENTRY's Experiment 5 is the first systematic empirical comparison of SHAP (over node/edge features), GNNExplainer (subgraph-level attribution), and PGExplainer (parametric, inductive, 10^8× faster) applied to agent interaction graphs, evaluated on three axes: explanation fidelity (faithfulness to model prediction via masked feature prediction flip), runtime overhead per explanation, and operator utility (structured questionnaire measuring simulatability, decomposability, and algorithmic transparency, n=12 professional software engineers). No prior work exists comparing these three methods in the agent graph domain. This is a direct contribution to the SHAP/XAI evaluation literature.

I also used SHAP in my published RAVA platform (DOI: 10.18080/jtde.v14n1.142) for network anomaly diagnosis — the only prior empirical work on SHAP in an operational AI deployment context.

I plan to apply for a DAAD Research Grant (deadline November 1, 2026), which requires a supervisor confirmation letter from a German professor. """ + COND + """

Would you be willing to discuss PhD opportunities and the possibility of providing a DAAD supervisor letter?

""" + SIG,
     "DAAD Research Grant\nGermany\nDeadline: Nov 1, 2026\nPortal: portal.daad.de\n€934/mo + health + travel\n\nNO IELTS at application",
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM CET\n📎 CV + SENTRY proposal\n💡 DAAD = you need supervisor letter by Oct\n🔁 Follow-up: +14 days"),

    ("🇩🇪 Germany\nProf. Lauritz Thamsen\nTU Berlin — AOT Group\n\n🎯 AI for distributed systems\nCloud anomaly detection\nML operations\n\n🔗 Scholarship: DAAD (Nov 1, 2026)",
     "lauritz.thamsen@tu-berlin.de",
     "PhD Inquiry 2026/27 — SENTRY: Distributed Runtime Monitoring for Multi-Agent LLM Systems | Scopus Q2 | DAAD Applicant",
     """Dear Prof. Thamsen,

Your research on AI for distributed systems and anomaly detection in cloud infrastructure is directly relevant to SENTRY's deployment context — my PhD proposal on non-invasive runtime monitoring for multi-agent LLM systems deployed in production cloud environments.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The distributed systems dimension of SENTRY is non-trivial: in production deployments, multi-agent LLM systems run as distributed cloud workloads — individual agents execute on separate containers, communicate via message queues, and share state through vector databases. SENTRY's monitoring framework must therefore aggregate execution events across distributed nodes, build a coherent interaction graph, and run the anomaly detector without introducing latency that disrupts the monitored system. This is precisely the kind of distributed observability problem your group studies in cloud systems. Extending SENTRY to a Kubernetes-native deployment model (sidecar containers, distributed trace aggregation) would be a significant systems contribution.

I also plan to apply for a DAAD Research Grant (deadline November 1, 2026), which requires a supervisor confirmation letter. """ + COND + """

Would you be open to discussing PhD opportunities and the possibility of serving as my DAAD supervisor?

""" + SIG,
     "DAAD Research Grant\nDeadline: Nov 1, 2026\nPortal: portal.daad.de\n€934/mo + health + travel\n\nNO IELTS at DAAD application",
     "📅 Send: Week of May 26, 2026\n⏰ 9:00 AM CET\n📎 CV + SENTRY proposal\n💡 Ask explicitly for DAAD supervisor letter\n🔁 Follow-up: +14 days"),

    ("🇳🇱 Netherlands\nProf. Mehdi Dastani\nUtrecht University\nFaculty of Science — Intelligent Systems\n\n⭐ EUROPE'S TOP multi-agent systems\nresearcher — perfect SENTRY match\n\n🔗 Scholarship: NWO grant / EU Horizon",
     "m.m.dastani@uu.nl",
     "PhD Inquiry 2026/27 — SENTRY: Runtime Behavioural Anomaly Detection for Multi-Agent LLM Coordination Failures | Scopus Q2",
     """Dear Prof. Dastani,

You are one of Europe's foremost researchers in multi-agent systems, and the failures that SENTRY addresses — inter-agent misalignment, coordination drift, and cascading delegation errors in deployed LLM systems — are precisely the failure modes your theoretical work on MAS coordination has characterised for classical agent architectures, now emerging at scale in LLM deployments.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The theoretical bridge is direct: SENTRY's three target failure categories — (i) agent reasoning loops (inter-agent misalignment), (ii) tool invocation anomalies (task verification failure), (iii) delegation drift (system design failure) — map onto the coordination failure taxonomy established in classical MAS research. What SENTRY adds is a data-driven detection mechanism for these failures at runtime in LLM deployments, where formal verification is intractable due to the non-deterministic, probabilistic execution model of LLM agents. Your expertise in MAS coordination semantics and formal agent programming would be invaluable for grounding SENTRY's failure taxonomy in established MAS theory — a contribution that would significantly strengthen the paper.

""" + PAPERS + """

The Netherlands does not require IELTS at the PhD application stage. """ + COND + """

Would you be open to discussing a potential PhD position in your group?

""" + SIG,
     "NWO (Netherlands Organisation\nfor Scientific Research)\nEU Horizon Europe\n\nNo specific scholarship\n— direct lab funding\nPhD salary: ~€2,400/mo\n(Dutch university employee)",
     "📅 Send: Week of May 19, 2026\n⏰ 9:00 AM CET\n📎 CV + SENTRY proposal\n💡 Netherlands = no IELTS at application\n🔁 Follow-up: +14 days\n⭐ BEST multi-agent systems match in EU"),

    ("🇬🇧 United Kingdom\nProf. Michael Fisher\nUniversity of Manchester\nSchool of Computer Science\n\n🎯 Formal verification of autonomous agents\nRuntime verification pioneer\n\n🔗 Scholarship: UKRI/EPSRC or self-funded",
     "Michael.Fisher@manchester.ac.uk",
     "PhD Inquiry 2026/27 — SENTRY: Empirical Runtime Monitoring as Complement to Formal Agent Verification | Scopus Q2",
     """Dear Prof. Fisher,

Your pioneering work on the formal verification of autonomous agent systems defines the theoretical space within which SENTRY — my PhD proposal on empirical runtime monitoring for multi-agent LLM systems — operates as a complementary, data-driven approach.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

The relationship between formal agent verification and SENTRY is complementary rather than competing. Formal methods (model checking, runtime verification of LTL/CTL specifications) provide guarantees for deterministic systems with finite, enumerable state spaces — they are challenged by the continuous, probabilistic, non-deterministic execution model of LLM agents. SENTRY operates in this gap: it uses a learned behavioural model (GAT trained on normal execution traces) to detect deviations empirically, without requiring a formal specification. This is not a replacement for formal methods — it is the monitoring approach that works when formal specification is intractable. I believe there is a compelling joint paper in comparing SENTRY's learned anomaly detection against specification-based runtime monitors on the same agent execution traces.

""" + PAPERS + """

I am preparing for IELTS (required for UK). """ + COND + """

Would you be open to discussing potential PhD opportunities?

""" + SIG,
     "UKRI / EPSRC PhD Studentship\n(competitive, university-specific)\n~£18,000/yr stipend\n\n⚠️ IELTS required for UK\nTarget: 6.5+\n\nAlternative: self-funded\nor DAAD-funded UK visit",
     "📅 Send: June 2026\n⚠️ UK requires IELTS before enrollment\n📎 CV + SENTRY proposal\n💡 Frame SENTRY vs formal verification\n🔁 Follow-up: +14 days"),

    ("🇸🇪 Sweden\nProf. Magnus Boman\nKTH Royal Institute of Technology\nSchool of EECS — Dept of CS\n\n🎯 Multi-agent AI, AI ethics + safety\nTrustworthy AI systems\n\n🔗 Scholarship: Swedish VR grant",
     "mab@kth.se",
     "PhD Inquiry 2026/27 — SENTRY: Trustworthy Runtime Monitoring for Multi-Agent LLM Systems | Scopus Q2 | 2026/27",
     """Dear Prof. Boman,

Your research on multi-agent AI and AI ethics/trustworthiness directly addresses the trustworthy AI objectives of SENTRY — my PhD proposal on runtime behavioural reliability monitoring for multi-agent LLM systems.

My name is Abd El Madjid Kahoul — MSc graduate from the University of Constantine 2, Algeria (2025, Top Class B).

""" + SENTRY_BRIEF + """

SENTRY's design principles — non-invasiveness, privacy-by-design (structural features only, no content capture), proportionality (graduated intervention), and auditability (immutable audit trail aligned with EU AI Act Article 12) — are trustworthy AI principles applied at the system monitoring layer. Your research on AI ethics and trustworthy agent systems would be directly relevant to evaluating and strengthening these principles, particularly the governance implications of automated intervention systems operating on deployed LLM agents in production environments.

""" + PAPERS + """

Sweden does not require IELTS at the PhD application stage. """ + COND + """

Would you have PhD openings in your group for 2026/2027?

""" + SIG,
     "Swedish Research Council (VR)\nVinnova (Sweden Innovation Agency)\nEU Horizon\n\nPhD salary: ~€2,300/mo\n(Swedish university employee)\nNo IELTS required",
     "📅 Send: June 2026\n✅ Sweden = no IELTS at application\n📎 CV + SENTRY proposal\n🔁 Follow-up: +14 days"),
]

for i, (country_prof, to, subj, body, sch, notes) in enumerate(europe_emails):
    r = 3 + i
    ws5.row_dimensions[r].height = 300
    bg = ALT[i % len(ALT)]
    C(ws5, r, 1, country_prof, bold=True, color="1F4E79", bg=bg, sz=9)
    C(ws5, r, 2, to, bold=True, color="117A65", bg=bg, sz=9)
    C(ws5, r, 3, subj, bold=True, color="6C3483", bg=bg, sz=9)
    C(ws5, r, 4, body, bg=bg, sz=9)
    C(ws5, r, 5, sch, bold=True, color="784212", bg=bg, sz=9)
    C(ws5, r, 6, notes, italic=True, color="E67E22", bg=bg, sz=9)

BD(ws5, 2, 2+len(europe_emails), 1, 6)
ws5.freeze_panes = "A3"
W(ws5, [26, 26, 34, 84, 20, 26])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SCHOLARSHIPS (exact steps, dates, documents)
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("💰 SCHOLARSHIPS GUIDE")
ws6.sheet_view.showGridLines = False
T(ws6, "SCHOLARSHIP APPLICATION GUIDE — Exact Steps · Exact Dates · Your Documents Filled In", "784212", 6)

for i, h in enumerate(["SCHOLARSHIP", "STATUS RIGHT NOW", "EXACT DEADLINE", "STEP-BY-STEP PROCESS", "YOUR DOCUMENTS (personalized)", "PORTAL + CONTACT"], 1):
    H(ws6, 2, i, h, bg="935116", sz=9)
ws6.row_dimensions[2].height = 28

schs = [
    ("1. ITALY PNRR PhD Positions\n\nUniversities:\n• Politecnico di Torino\n• Politecnico di Milano\n• Univ. Bologna\n• Univ. Bolzano\n• Univ. Pisa\n\nFunding: €1,200/mo net\n3 years, employee status",
     "🟡 UPCOMING — URGENT\nCalls expected: May–June 2026\n\n✅ NO IELTS required\n✅ English transcripts ready\n✅ SENTRY proposal ready\n✅ Algeria eligible\n\nCheck portals DAILY:",
     "📅 VARIABLE\nCalls open: May–June 2026\nClose: typically 4–6 weeks\nafter opening\n\n⚠️ CHECK NOW:\ndottorato.polito.it\nphd.polimi.it\ndottorato.unibo.it",
     """STEP 1 — EMAIL professors first (see Sheet ITALY):
→ Chiasserini (PoliTo): chiasserini@polito.it
→ Montali (Bolzano): marco.montali@unibz.it
→ Ardagna (PoliMi): danilo.ardagna@polimi.it

STEP 2 — Check portals DAILY for call publication:
→ dottorato.polito.it
→ phd.polimi.it
→ dottorato.unibo.it
→ phd.unibz.it

STEP 3 — When call is published:
→ Download the "bando" (call document)
→ Find the position matching your area (AI / trustworthy AI / distributed systems)
→ Note the specific position number

STEP 4 — Register on the portal + upload:
→ SENTRY_Proposal_v2_1.pdf as research proposal ✅
→ CV (2 pages)
→ English transcripts ✅
→ 2 reference letters
→ Passport copy

STEP 5 — Selection:
→ Thesis proposal defense (15 min in English)
→ Prepare: research presentation

STEP 6 — Results:
→ 4–6 weeks after call closes
→ If awarded: PNRR contract starts Oct 2026 or Jan 2027""",
     """✅ RESEARCH PROPOSAL:
SENTRY_Proposal_v2_1.pdf — READY

✅ ACADEMIC CV (2 pages):
• Abd El Madjid Kahoul
• MSc SE & IS, Constantine 2 (2025, Top Class B)
• BSc Software Engineering (2023, Class A)
• Publications: 2 Scopus (DOIs listed)
• Ooredoo Algeria AI Intern (Apr–Jul 2025)
• Skills: Python, TF, SHAP, GNN, Django, AWS

✅ TRANSCRIPTS (English — already have!):
• MSc all semesters
• BSc all semesters

✅ DIPLOMAS: certified copies

✅ 2 REFERENCE LETTERS:
• MSc thesis supervisor (Constantine 2)
• Ooredoo Algeria internship supervisor
• On official letterhead, signed + stamp

✅ PASSPORT COPY

❌ NO IELTS REQUIRED — EVER in Italy""",
     "🌐 dottorato.polito.it\n🌐 phd.polimi.it\n🌐 dottorato.unibo.it\n🌐 phd.unibz.it\n\n📧 PoliTo PhD office:\ndottorato@polito.it\n\n📧 PoliMi PhD office:\nphd-programme@polimi.it\n\n💡 Also check EURAXESS\nfor Italian positions:\neuraxess.ec.europa.eu"),

    ("2. JAPAN MEXT\nEmbassy Route\n\n¥145,000/mo (~€900)\n+ full tuition waiver\n+ round-trip airfare\n\nStart as 'research student'\n→ convert to PhD",
     "🔴 OPEN RIGHT NOW\nEmbassy route deadline:\nMay–June 2026\n\n⚠️ CONTACT EMBASSY\nTHIS WEEK\n\n✅ NO IELTS required\n✅ English research plan OK\n✅ Algeria eligible",
     "📅 URGENT\nEmbassy route:\nMay–June 2026\n(confirm exact date\nwith embassy)\n\n📍 Japanese Embassy\nAlgiers:\n25 Chemin Gadouche\nHydra, Algiers\n+213 21 54 04 08",
     """STEP 1 — THIS WEEK: Visit or call Japanese Embassy:
→ Address: 25 Chemin Gadouche, Hydra, Algiers
→ Phone: +213 21 54 04 08
→ Email: jp-emb-algiers@ws.mofa.go.jp
→ Say: "I want to apply for the 2026 MEXT Research Scholarship — embassy route"
→ Get: application forms + EXACT current deadline

STEP 2 — Download MEXT forms:
→ mext.go.jp/en/policy/education/highered/title02/detail02/sdetail02/1373897.htm
→ Fill: MEXT application form (personal info)

STEP 3 — Write Research Plan (2 pages, English):
→ Title: "SENTRY: Runtime Behavioural Reliability Monitoring for Multi-Agent LLM Systems"
→ Based on SENTRY_Proposal_v2_1.pdf — summarize in 2 pages
→ Name proposed supervisor (e.g., Prof. Matsuo, UTokyo)
→ Explain why Japan: "Japan leads globally in LLM research (Matsuo Lab) and AI policy (AI Strategy)"

STEP 4 — Health certificate:
→ Download health form from embassy
→ Complete at hospital in Constantine or Algiers

STEP 5 — Assemble package:
→ Completed MEXT form
→ Research plan (2 pages)
→ English transcripts ✅
→ Diplomas (certified copies)
→ 2 recommendation letters
→ Health certificate
→ Photos (passport-style)
→ Passport copy

STEP 6 — Submit to embassy by their deadline
STEP 7 — Embassy interview (in English)
STEP 8 — If selected: embassy sends to MEXT → university acceptance → departure""",
     """✅ MEXT APPLICATION FORM:
Download from embassy or mext.go.jp

✅ RESEARCH PLAN (2 pages — WRITE NOW):
Based on SENTRY proposal — summarize:
• Your background (MSc + 2 Scopus papers)
• The problem (silent failures in multi-agent LLMs)
• Your approach (GAT + GNNExplainer + intervention)
• Why Japan and this professor

✅ TRANSCRIPTS (English — already have!):
• MSc all semesters
• BSc all semesters

✅ DIPLOMAS (certified copies):
• MSc diploma (2025)
• BSc diploma (2023)

✅ 2 RECOMMENDATION LETTERS:
• MSc thesis supervisor
• Ooredoo Algeria manager

✅ HEALTH CERTIFICATE:
Download form from embassy → fill at hospital

✅ PASSPORT COPY

✅ PHOTOS (passport style, 3×4cm)

❌ NO IELTS — ever for MEXT""",
     "📍 Japanese Embassy Algiers:\n25 Chemin Gadouche\nHydra, Algiers\n+213 21 54 04 08\njp-emb-algiers@ws.mofa.go.jp\n\n🌐 MEXT info:\nmext.go.jp/en\n\n🌐 UTokyo international:\nu-tokyo.ac.jp/en/admissions\n\n🌐 RIKEN:\nriken.jp/en"),

    ("3. MALAYSIA MIS\n(Malaysia International Scholarship)\n\nRM 3,000/mo (~€600)\n+ full tuition\n+ health insurance\n+ round-trip airfare\n\nDuration: up to 4 years",
     "⚠️ CHECK NOW\nscholarship.moe.gov.my\n\nTypical cycle: March–June\nMight be open RIGHT NOW\n\nIf closed 2026:\nnext cycle March 2027\n\n✅ NO IELTS required\n✅ Algeria eligible",
     "📅 URGENT — CHECK:\nscholarship.moe.gov.my\nMay/June deadline possible\n\nIf closed:\nMarch 2027 (next cycle)\n\n⚠️ Still email professors\nNOW to prepare",
     """STEP 1 — RIGHT NOW:
→ Go to scholarship.moe.gov.my
→ Check: is the 2026 MIS cycle open?
→ If YES: proceed to Step 2 immediately
→ If NO: still email professors (build relationship for 2027)

STEP 2 — Email professors (see Sheet MALAYSIA):
→ Prof. Ahmad (UTM): shams@utm.my
→ Prof. Shahrul (UKM): shahrul@ukm.edu.my
→ Prof. Varathan (UM): kasturi@um.edu.my
→ Goal: get supervisor agreement email

STEP 3 — Complete MIS online application:
→ Proposed university: UTM or UM or UKM
→ Proposed supervisor: from step 2
→ Research area: Trustworthy AI / Multi-Agent Systems

STEP 4 — Upload documents:
→ MIS application form (online)
→ English transcripts ✅
→ Diplomas (certified)
→ Personal statement (adapt SENTRY proposal)
→ 2 reference letters
→ CV
→ Passport copy

STEP 5 — MIS evaluation + interview
STEP 6 — If awarded: Malaysian student visa + enrollment""",
     """✅ MIS ONLINE APPLICATION:
scholarship.moe.gov.my

✅ PERSONAL STATEMENT (2-3 pages):
Adapt SENTRY proposal for Malaysia context:
"Trustworthy AI for Southeast Asian Digital Economy"
Focus on: AI governance, regional tech development

✅ TRANSCRIPTS (English — already have!)

✅ DIPLOMAS (certified copies)

✅ CV (2 pages):
Standard academic CV with publications

✅ 2 REFERENCE LETTERS:
MSc supervisor + Ooredoo manager

✅ PASSPORT COPY

NO IELTS REQUIRED FOR MIS""",
     "🌐 MIS portal:\nscholarship.moe.gov.my\n\n📧 MIS secretariat:\nmis@moe.gov.my\n\n🌐 UTM international:\nutmgraduate.utm.my\n\n🌐 UM international:\num.edu.my/main/research\n\n📧 UTM admission:\nadmission.graduate@utm.my"),

    ("4. VLIR-UOS Belgium\n\nFull funding:\n€900/mo + tuition\n+ flight + insurance\n\nLinked universities:\nGhent, KU Leuven\nUCLouvain\n\nKey target:\nProf. De Turck (IDLab)",
     "🟡 UPCOMING\nOpens: ~November 2026\nDeadline: ~March 31, 2027\n\n✅ Algeria eligible\n✅ NO IELTS required\n✅ De Turck named in\nyour SENTRY proposal\n\nContact De Turck NOW\nto build the relationship",
     "📅 Opens: ~Nov 2026\n⏰ Deadline: ~Mar 31, 2027\n\nVerify at: vliruos.be\n\n⚠️ Must contact De Turck\nNOW — supervisor\nagreement = +40%\nacceptance rate",
     """STEP 1 — NOW: Email De Turck (Sheet EUROPE email #1):
→ Filip.DeTurck@ugent.be
→ Mention IDLab is named in your SENTRY proposal
→ Goal: informal research interest

STEP 2 — June–October 2026:
→ Maintain email contact with De Turck
→ Share updates on SENTRY (any new work)
→ Get "I'm interested in supervising you" statement

STEP 3 — November 2026:
→ vliruos.be/en/scholarships/22
→ Create VLIR-UOS account
→ Begin application form

STEP 4 — January–March 2027:
→ Complete and submit application
→ De Turck confirms supervision in VLIR system
→ Upload all documents

STEP 5 — Results: April–May 2027
STEP 6 — If awarded: Belgium visa → Ghent Oct 2027""",
     """✅ VLIR APPLICATION FORM (vliruos.be)

✅ ACADEMIC CV (max 4 pages):
• Education: MSc (2025, Top B), BSc (2023, A)
• 2 Scopus publications with DOIs
• Ooredoo Algeria internship
• SENTRY proposal summary (1 para)

✅ MOTIVATION LETTER (1 page):
See Sheet MOTIVATION LETTERS (VLIR version)

✅ RESEARCH PROPOSAL (2-3 pages):
SENTRY condensed version for VLIR
Focus: AI reliability monitoring for distributed systems

✅ TRANSCRIPTS (English ✅)

✅ DIPLOMAS (certified copies)

✅ 2 REFERENCE LETTERS:
MSc supervisor + Ooredoo manager

✅ SUPERVISOR AGREEMENT:
Email from De Turck confirming interest
(He uploads this in VLIR system himself)

❌ NO IELTS — NEVER required for VLIR""",
     "🌐 vliruos.be/en/scholarships/22\n📋 portal: scholarship.vliruos.be\n📧 info@vliruos.be\n\n📍 VLIR-UOS:\nEgmontstraat 11\n1000 Brussels, Belgium\n\n🌐 Ghent PhD portal:\nphd.ugent.be"),

    ("5. DAAD Research Grant\nGermany\n\n€934/mo + health insurance\n+ travel allowance\n\nLinked universities:\nTU Berlin, TU Munich\nUniv. Göttingen, KIT\n\nTarget supervisors:\nMüller, Thamsen",
     "⚠️ UPCOMING\nApplication: Oct–Nov 2026\n\n⚠️ SUPERVISOR LETTER\nrequired — contact\nGerman professors NOW\n\n✅ NO IELTS at application\n✅ Algeria eligible",
     "📅 HARD DEADLINE:\n~November 1, 2026\n\n⚠️ CANNOT submit\nwithout supervisor letter\n\nContact German professors\nNOW — you need their\nagreement by Sept 2026",
     """STEP 1 — NOW: Email German professors:
→ Müller (TU Berlin): use Sheet EUROPE email #4
→ Thamsen (TU Berlin): use Sheet EUROPE email #5
→ Ask explicitly: "Would you be willing to serve as DAAD supervisor?"

STEP 2 — If professor agrees:
→ Ask for written confirmation (email is enough initially)
→ Request formal supervisor letter for October

STEP 3 — September 2026:
→ portal.daad.de → create account
→ Select: "Research Grants — One Year for Doctoral Candidates" (code 57214057)
→ Begin filling application

STEP 4 — October 2026:
→ Write 5-page research proposal (SENTRY extended)
→ Get formal supervisor letter from German professor
→ Upload all documents to DAAD portal

STEP 5 — November 1: SUBMIT (hard deadline)
STEP 6 — Results: February–March 2027""",
     """✅ DAAD APPLICATION FORM (portal.daad.de)
Grant code: 57214057

✅ RESEARCH PROPOSAL (5 pages):
"SENTRY: Runtime Graph-Based Monitoring for
Multi-Agent LLM Systems"
Extend SENTRY proposal with:
• Review of Cemri et al., Dong et al., Huang et al.
• German university specific direction
• 12-month timeline at TU Berlin/TUM

✅ SUPERVISOR LETTER (CRITICAL):
→ From German professor on university letterhead
→ Must confirm: supervision willingness + topic + lab

✅ CV (3 pages — DAAD format)

✅ TRANSCRIPTS (English ✅)

✅ DIPLOMAS

✅ 2 REFERENCE LETTERS:
MSc supervisor + Ooredoo manager

✅ LANGUAGE CERT (optional at application)
→ Include IELTS if you have it by Nov

❌ NO IELTS required to APPLY""",
     "🌐 daad.de/en\n📋 portal.daad.de\n📌 Grant code: 57214057\n\n📧 DAAD Algeria:\nalgiers@daad.de\n\n📍 DAAD Algiers:\n12 Rue de Tripoli\nHussein Dey, Algiers\n+213 23 77 57 62\n(VISIT IN PERSON!)"),

    ("6. MSCA Doctoral Networks\nEU Horizon Europe\n\n~€3,000–3,200/mo\n(HIGHEST FUNDING AVAILABLE)\n+ mobility + family allowance\n\nAny EU consortium\nwith matching topic",
     "✅ ROLLING — Check WEEKLY\nnew positions open\nyear-round\n\n⚠️ Set EURAXESS alert\nTODAY\n\nKeywords to use:\n• 'trustworthy AI'\n• 'multi-agent systems'\n• 'graph neural networks'\n• 'LLM reliability'",
     "📅 ROLLING\nEach network has\nits own deadline\n\nTypically 4–8 weeks\nafter position posted\n\n⚠️ Apply immediately\nwhen you find a match",
     """STEP 1 — TODAY:
→ euraxess.ec.europa.eu/jobs/search
→ Search: "trustworthy AI" + "doctoral"
→ Search: "multi-agent systems" + "doctoral"
→ Search: "graph neural networks" + "PhD"
→ Click "Create Alert" for each search

STEP 2 — When matching ESR position found:
→ Read description carefully
→ Google the PI — check their research
→ Email the PI BEFORE applying (use SENTRY introduction)
→ Express specific interest in their position

STEP 3 — Apply to specific position:
→ Each network has own portal
→ Most require: Europass CV + cover letter + transcript + 2 refs

STEP 4 — Selection:
→ Written application → video interview → committee
→ Timeline: 4–8 weeks after deadline

STEP 5 — If selected:
→ Employment contract (not scholarship)
→ €3,000–3,200/mo depending on country""",
     """✅ EUROPASS CV (mandatory format):
europa.eu/europass
• Both Scopus publications with DOIs
• SENTRY proposal summary
• Skills: Python, TF, GNN, SHAP, AWS

✅ COVER LETTER (1 page):
Specific to each ESR position
Reference SENTRY components that match the project

✅ TRANSCRIPTS (English ✅)

✅ 2 REFERENCE LETTERS

✅ PASSPORT COPY

❌ NO IELTS at most MSCA networks""",
     "🌐 euraxess.ec.europa.eu/jobs/search\n\n📋 Europass CV:\neuropass.eu.europa.eu\n\n🔍 Weekly alerts for:\n• 'trustworthy AI'\n• 'multi-agent systems'\n• 'LLM reliability'\n• 'graph neural networks'\n\n💡 Also: findaphd.com"),

    ("7. CSC — Chinese Govt Scholarship\nChina\n\nFull tuition\n+ CNY 3,500/mo\n+ free housing\n\nTarget unis:\nTsinghua, PKU\nShanghai AI Lab\n\nField: AI Safety / LLM",
     "⚠️ UPCOMING\nEmail Chinese professors\nNOW to build relationship\n\nCSC portal opens:\nFeb–Apr 2027\n\n✅ NO IELTS required\n✅ Algeria eligible\n~70-80% accept rate",
     "📅 CSC PORTAL:\nFeb–Apr 2027\n\nContact professors NOW\n(July 2026) to get\npre-acceptance letter\nfor CSC application",
     """STEP 1 — July 2026: Email Chinese professor:
→ Prof. Jie Tang (Tsinghua, GLM/ChatGLM creator):
  tangjie@tsinghua.edu.cn
→ Prof. Hang Li (ByteDance Research / SJTU):
  hangli.hl@bytedance.com
→ Introduce SENTRY + your 2 Scopus papers
→ Goal: get pre-acceptance email

STEP 2 — February 2027 (CSC portal opens):
→ studyinchina.csc.edu.cn → create account
→ Fill application: name supervisor, research area (AI Safety)

STEP 3 — Feb–April 2027:
→ Get university admission notice + JW202 form from BUPT/Tsinghua
→ Upload all documents to CSC portal

STEP 4 — May 2027: Results announced""",
     """✅ CSC APPLICATION (studyinchina.csc.edu.cn)

✅ STUDY PLAN (2 pages):
"SENTRY at Tsinghua: Foundation Model Safety Monitoring"
Connect to Chinese AI policy (AI governance, safety)

✅ UNIVERSITY ADMISSION NOTICE

✅ JW202 FORM (from university)

✅ TRANSCRIPTS (English ✅)

✅ DIPLOMAS (certified, notarized)

✅ HEALTH CERTIFICATE

✅ 2 REFERENCE LETTERS

✅ PASSPORT COPY

❌ NO IELTS for CSC""",
     "🌐 studyinchina.csc.edu.cn\n\n🌐 Tsinghua AI:\nai.tsinghua.edu.cn\n\n📧 Prof. Jie Tang:\ntangjie@tsinghua.edu.cn\n(GLM/ChatGLM creator —\nLLM systems expert)\n\n📍 Chinese Embassy Algeria:\n7 Ch. Prince d'Annam\nEl Biar, Algiers"),
]

for i, (name, status, deadline, steps, docs, portal) in enumerate(schs):
    r = 3 + i
    ws6.row_dimensions[r].height = 210
    bg = ALT[i % len(ALT)]
    C(ws6, r, 1, name, bold=True, color="784212", bg=bg, sz=9)
    C(ws6, r, 2, status, bold=True, color="117A65", bg=bg, sz=9)
    C(ws6, r, 3, deadline, bold=True, color="E74C3C", bg=bg, sz=9)
    C(ws6, r, 4, steps, bg=bg, sz=9)
    C(ws6, r, 5, docs, bg=bg, sz=9)
    C(ws6, r, 6, portal, italic=True, color="2471A3", bg=bg, sz=9)

BD(ws6, 2, 2+len(schs), 1, 6)
ws6.freeze_panes = "A3"
W(ws6, [22, 18, 16, 54, 46, 22])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 — MOTIVATION LETTERS
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("📝 MOTIVATION LETTERS")
ws7.sheet_view.showGridLines = False
T(ws7, "MOTIVATION LETTERS — 4 Complete Letters Using Your CV + SENTRY Proposal | Print → Sign → Submit", "6C3483", 3)

for i, h in enumerate(["TARGET (scholarship / university)", "COMPLETE LETTER TEXT", "FORMATTING INSTRUCTIONS"], 1):
    H(ws7, 2, i, h, bg="7D3C98", sz=10)
ws7.row_dimensions[2].height = 28

letters = [
    ("VLIR-UOS Scholarship\n(Belgium — Ghent University)\n\nMax: 1 page\nLanguage: English\nFont: Arial 11pt\nMargins: 2cm",
     """MOTIVATION LETTER — VLIR-UOS SCHOLARSHIP APPLICATION
PhD at Ghent University — IDLab | AI for Trustworthy Multi-Agent Systems

Abd El Madjid Kahoul
Constantine, Algeria | abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17

[Date]

To the VLIR-UOS Scholarship Selection Committee,

I am writing to apply for the VLIR-UOS scholarship to pursue a fully funded PhD at Ghent University's IDLab under the supervision of Prof. Filip De Turck. My PhD research proposal, SENTRY — A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems in Operational Environments — directly extends IDLab's established expertise in distributed system observability and AI-driven network management to the LLM agent execution layer. IDLab is specifically named in my SENTRY proposal as the primary target laboratory for this research.

ACADEMIC AND RESEARCH BACKGROUND. I hold a Master's degree in Software Engineering and Intelligent Systems from the University of Constantine 2, Algeria (2025, Top Class B), preceded by a Bachelor's degree in Software Engineering (2023, Class A). My academic work has produced two Scopus-indexed publications — rare at the Master's level — which provide the methodological foundation for SENTRY:

(1) As first author: "Hybrid Deep Learning Ensemble with Dynamic Fusion for Reliable Anomaly Detection in Operational LTE RANs" — JTDE, Scopus Q2, 2025. DOI: 10.18080/jtde.v13n4.1363. This paper demonstrates my ability to design and validate anomaly detection systems on real production data — the same capability SENTRY applies to LLM agent execution traces.

(2) As co-author: "RAVA: A Human-in-the-Loop Hybrid Platform for Explainable Anomaly Detection" — JTDE, 2026. DOI: 10.18080/jtde.v14n1.142. RAVA's SHAP-based explainability framework directly informs SENTRY's GNNExplainer/SHAP comparison in the agent graph domain.

THE SENTRY PROPOSAL. Multi-agent LLM systems now underpin commercial AIOps, NOC, and software development pipelines — yet empirical studies document task success rates as low as 25%, with failures that are structural and silent: no exception is raised when a reasoning loop consumes unbounded tokens or an agent's behaviour drifts from its objective. SENTRY addresses this gap with a non-invasive monitoring framework: callback hooks capture agent execution events without code modification, these events are assembled into dynamic interaction graphs, and a Graph Attention Network detects structural deviations from learned normal behaviour. A graduated intervention engine responds proportionally, and SHAP/GNNExplainer explainability provides operator-readable justifications.

WHY GHENT AND IDLab. IDLab's work on runtime observability for distributed application management (Sebrechts et al., Int. J. Network Management, 2018) and its EU Horizon projects on autonomous system management create the ideal environment for SENTRY's development. The transition from network telemetry observability to LLM agent execution observability is a natural extension of IDLab's research agenda.

WHY VLIR-UOS. The VLIR-UOS scholarship reflects a commitment to developing research capacity in the Global South that I am committed to serving. As an Algerian researcher with Scopus-level publications and a concrete, fundable research proposal, I am ready to make an immediate contribution to IDLab's research output. I am committed to returning knowledge and expertise to Algeria's growing AI and tech sector.

Respectfully,
Abd El Madjid Kahoul""",
     "✅ Replace [Date] with the actual date\n✅ Print on white A4, no borders or colour\n✅ Sign in blue or black pen\n✅ Scan as PDF\n✅ Keep to exactly 1 page (Arial 11pt, 2cm margins)\n✅ Submit with all other VLIR documents"),

    ("DAAD Research Grant\n(Germany — TU Berlin / TU Munich)\n\nMax: 2 pages\nLanguage: English\nFont: Arial 11pt",
     """MOTIVATION LETTER — DAAD RESEARCH GRANT APPLICATION
PhD in Trustworthy AI and Multi-Agent System Reliability

Abd El Madjid Kahoul
Constantine, Algeria | abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17

[Date]

To the DAAD Selection Committee,

I am applying for the DAAD Research Grant to pursue a PhD at TU Berlin's Machine Learning Group (Prof. Klaus-Robert Müller) on the explainability component of SENTRY — A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems. SHAP, co-developed by Prof. Müller, is one of three explanation methods SENTRY will compare empirically in a novel domain — agent interaction graphs — where no prior comparison exists.

ACADEMIC AND RESEARCH BACKGROUND. I hold a Master's in Software Engineering and Intelligent Systems from the University of Constantine 2, Algeria (2025, Top Class B) and a Bachelor's in Software Engineering (2023, Class A). My research has produced two Scopus publications during my Master's:

"Hybrid Deep Learning Ensemble with Dynamic Fusion for Reliable Anomaly Detection in Operational LTE RANs" — JTDE, Scopus Q2, 2025. DOI: 10.18080/jtde.v13n4.1363.
"RAVA: A Human-in-the-Loop Hybrid Platform for Explainable Anomaly Detection and Diagnosis in LTE RANs" — JTDE, 2026. DOI: 10.18080/jtde.v14n1.142.

Both were validated on real production data from Ooredoo Algeria, where I also completed an AI Developer internship (April–July 2025). RAVA's use of SHAP for operational anomaly diagnosis is the direct precursor to SENTRY's explainability module.

RESEARCH PROPOSAL. Multi-agent LLM systems deployed in production fail silently — reasoning loops, tool-call anomalies, and inter-agent delegation drift produce no observable error signal. SENTRY addresses this with: (1) a non-invasive trace collector hooking into LangGraph/AutoGen event buses, (2) a dynamic agent interaction graph built from execution events, (3) a Graph Attention Network detecting structural deviations from learned normal behaviour, (4) a graduated intervention engine, and (5) comparative XAI evaluation of SHAP, GNNExplainer, and PGExplainer on agent graphs.

The DAAD-funded year at TU Berlin would focus on Objective 6: the first systematic empirical comparison of SHAP over node/edge features, GNNExplainer (subgraph attribution), and PGExplainer (parametric, 10^8× faster at runtime) on agent interaction graphs, evaluated on explanation fidelity (prediction flip on masked features), runtime overhead, and operator utility (n=12 professional software engineers, measuring simulatability, decomposability, and algorithmic transparency). This is a direct contribution to the SHAP/XAI evaluation literature and to the emerging field of trustworthy multi-agent AI.

TECHNICAL QUALIFICATIONS. Python, TensorFlow, PyTorch Geometric, SHAP, GNNExplainer, Scikit-learn, LangGraph integration, AWS Cloud Practitioner (certified 2024).

WHY GERMANY AND DAAD. Germany's EU AI Act leadership and TU Berlin's position as a top technical university in Europe make this the ideal environment for this research. DAAD's support for international research talent represents the kind of investment in global scientific exchange that I am committed to returning through publications, open-source tools, and knowledge transfer to Algeria's AI research community.

Respectfully,
Abd El Madjid Kahoul""",
     "✅ Replace Prof. Müller with your actual German supervisor if different\n✅ Replace TU Berlin with actual university\n✅ Print and sign before scanning\n✅ This is 2 pages — acceptable for DAAD\n✅ Must accompany the 5-page research proposal (separate document)\n✅ Upload to portal.daad.de"),

    ("Italy PNRR PhD Call\n(Any Italian university)\n\nMax: 1 page\nLanguage: English\nFont: Arial 11pt\nNo IELTS required",
     """COVER LETTER — PhD POSITION APPLICATION
Politecnico di Torino | NetAI Lab | Funded PNRR PhD Position

Abd El Madjid Kahoul
Constantine, Algeria | abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17

[Date]

Dear Prof. Chiasserini / Dear Selection Committee,

I am writing to apply for a PNRR-funded PhD position in the NetAI Lab at the Politecnico di Torino. My PhD research proposal, SENTRY — A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems — creates a direct research bridge to your group's work on LLM-driven NOC automation, which I specifically mention in the SENTRY proposal as a target validation domain and co-publication opportunity.

RESEARCH PUBLICATIONS. As first author, I published "Hybrid Deep Learning Ensemble with Dynamic Fusion for Reliable Anomaly Detection in Operational LTE RANs" (JTDE, Scopus Q2, 2025; DOI: 10.18080/jtde.v13n4.1363) — a hybrid DL anomaly detection system validated on real production LTE data from Ooredoo Algeria. My co-authored second paper introduces RAVA, a SHAP-based XAI platform for interactive operator diagnosis (JTDE, 2026; DOI: 10.18080/jtde.v14n1.142). These publications demonstrate my ability to build, validate, and deploy AI systems on real operational data — the same capability required by SENTRY.

THE SENTRY RESEARCH. SENTRY is a non-invasive runtime monitoring framework for multi-agent LLM systems: callback hooks capture agent execution events without modifying agent code; events are assembled into dynamic interaction graphs; a Graph Attention Network detects structural deviations; a graduated intervention engine responds (Monitor → Pause → Rollback → Handoff); and SHAP/GNNExplainer/PGExplainer comparison provides operator-readable justifications.

CONNECTION TO NETAI LAB. Your group's work on LLM-driven NOC automation is precisely the operational deployment context for SENTRY's NOC simulation validation (stretch goal). A SENTRY-monitored multi-agent NOC pipeline — where the monitoring framework detects when individual NOC agents enter reasoning loops or deviate from expected delegation patterns — is a direct co-publication opportunity.

INDUSTRY EXPERIENCE. AI Developer intern at Ooredoo Algeria (April–July 2025): ML/DL models for anomaly detection in production LTE KPI streams.

I am available for a thesis proposal defense at any time. Italy requires no language certificate at the application stage.

Respectfully,
Abd El Madjid Kahoul""",
     "✅ Replace 'Chiasserini' with actual professor name\n✅ Replace 'NetAI Lab' with correct lab name\n✅ Replace PoliTo with actual Italian university if different\n✅ Print → sign → scan as PDF\n✅ Attach to PNRR portal application along with SENTRY proposal PDF\n✅ NO IELTS required — do not mention language"),

    ("General EU Cover Letter\n(For any EU professor email)\n\nMax: 1 page\nLanguage: English\nAdapt for each professor",
     """COVER LETTER — PhD POSITION INQUIRY

Abd El Madjid Kahoul
Constantine, Algeria | abdelmadjid.kahoul@hotmail.com | +213 658 63 99 17 | LinkedIn: kahoul-abd-el-madjid

[Date]

Dear Prof. [NAME],

I am writing to inquire about a funded PhD position in your group for 2026/2027. I am a Master's graduate in Software Engineering and Intelligent Systems from the University of Constantine 2, Algeria (2025, Top Class B), with two Scopus-indexed publications and a concrete, fully developed PhD research proposal directly relevant to your group's research agenda.

RESEARCH PROPOSAL. My PhD proposal, SENTRY — A Runtime Behavioural Reliability Monitor for Multi-Agent LLM Systems in Operational Environments — addresses a critical unsolved reliability problem in deployed multi-agent AI systems. SENTRY hooks into agent framework event buses (LangGraph, AutoGen) without modifying agent code, constructs dynamic inter-agent interaction graphs, and applies a Graph Attention Network (GAT) to detect structural behavioural anomalies. A graduated intervention engine (Monitor → Pause → Rollback → Handoff) responds proportionally, and SHAP/GNNExplainer/PGExplainer explainability produces operator-readable justifications for each alert. The proposal is complete and available on request.

PUBLICATIONS. Two Scopus-indexed publications from my Master's studies:
(1) First-author: "Hybrid Deep Learning Ensemble with Dynamic Fusion for Reliable Anomaly Detection in Operational LTE RANs" — JTDE Scopus Q2, 2025. DOI: 10.18080/jtde.v13n4.1363
(2) Co-author: "RAVA: A Human-in-the-Loop Hybrid Platform for Explainable Anomaly Detection and Diagnosis in LTE RANs" — JTDE, 2026. DOI: 10.18080/jtde.v14n1.142

INDUSTRY EXPERIENCE. AI Developer intern at Ooredoo Algeria (April–July 2025) — ML/DL for anomaly detection in production LTE KPI data.

SKILLS. Python · TensorFlow · PyTorch Geometric · SHAP · GNNExplainer · Scikit-learn · LangGraph · Docker · AWS Certified (2024)

LANGUAGES. Arabic (native) · French (TCF B2) · English (working proficiency — demonstrated through international publications)

[ADD ONE PARAGRAPH HERE: "I have read your work on [SPECIFIC PAPER/PROJECT] and believe [SPECIFIC CONNECTION TO SENTRY]"]

I am available for a research interview at any time to demonstrate my English proficiency and technical depth. I would warmly welcome a conditional acceptance arrangement pending an IELTS certificate before enrollment.

Please find my CV and SENTRY proposal attached.

Sincerely,
Abd El Madjid Kahoul""",
     "✅ Replace [Date] with actual date\n✅ Replace [NAME] with professor name\n✅ ALWAYS add the personalized paragraph — this is the most important part\n✅ Keep to exactly 1 page\n✅ Print → sign → scan as PDF\n✅ Attach alongside CV to EVERY application"),
]

for i, (target, letter, notes) in enumerate(letters):
    r = 3 + i
    ws7.row_dimensions[r].height = 400
    bg = ALT[i % len(ALT)]
    C(ws7, r, 1, target, bold=True, color="6C3483", bg=bg, sz=9)
    C(ws7, r, 2, letter, bg=bg, sz=9)
    C(ws7, r, 3, notes, italic=True, color="117A65", bg=bg, sz=9)

BD(ws7, 2, 2+len(letters), 1, 3)
ws7.freeze_panes = "A3"
W(ws7, [22, 108, 32])

# Save
out = "C:\\Users\\abdel\\Downloads\\Task Management Website Design\\SENTRY_PhD_System_Kahoul.xlsx"
wb.save(out)
print("SAVED:", out)
